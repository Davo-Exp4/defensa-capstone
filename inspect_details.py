import openpyxl
from collections import Counter
import re

wb_crudo = openpyxl.load_workbook("data/cohorte_pasada_crudo.xlsx", data_only=True)
sheet = wb_crudo.active

headers = [cell.value for cell in sheet[1]]
print("Headers length:", len(headers))

# Row data structures
# Column 15 is 'Seleccione el nombre del Estudiante'
# Column 12 is 'Seleccione su nombre (Evaluador)'
# Column 18 is 'Apertura: problema y objetivo' (index 17 in python 0-indexed list)
# Column 21 is 'Metodología (nivel adecuado)' (index 20 in python)
# Column 24 is 'Resultados y evidencia' (index 23 in python)
# Column 27 is 'Coherencia y manejo del tiempo' (index 26 in python)
# Column 30 is 'Diapositivas como apoyo' (index 29 in python)
# Column 33 is 'Respuestas a preguntas' (index 32 in python)
# Column 36 is 'Cierre: aporte a próximos pasos' (index 35 in python)

students_raw = []
evaluators_raw = []
evals_per_student = Counter()
student_to_evaluators = {}

criterias_cols = [17, 20, 23, 26, 29, 32, 35] # 0-indexed: 18th, 21st, 24th, 27th, 30th, 33rd, 36th columns

criterias_options = {h: Counter() for h in [headers[i] for i in criterias_cols]}

for r_idx in range(2, sheet.max_row + 1):
    row_vals = [sheet.cell(row=r_idx, column=c).value for c in range(1, len(headers) + 1)]
    student = row_vals[14] # 'Seleccione el nombre del Estudiante'
    evaluator = row_vals[11] # 'Seleccione su nombre (Evaluador)'
    students_raw.append(student)
    evaluators_raw.append(evaluator)
    if student:
        evals_per_student[student] += 1
        student_to_evaluators.setdefault(student, []).append(evaluator)
    
    for c_idx in criterias_cols:
        val = row_vals[c_idx]
        if val is not None:
            criterias_options[headers[c_idx]][val] += 1

print("\n--- Student Stats ---")
print(f"Total rows in raw (excluding header): {len(students_raw)}")
print(f"Number of unique student names in raw: {len(set(students_raw))}")
print(f"Number of unique evaluators: {len(set(evaluators_raw))}")

print("\n--- Distribution of evaluations per student ---")
print(Counter(evals_per_student.values()))

print("\n--- Sample of students with other than 3 evaluations ---")
for s, count in evals_per_student.items():
    if count != 3:
        print(f"Student: '{s}', Evals count: {count}, Evaluators: {student_to_evaluators[s]}")

print("\n--- Unique values for each rubric criteria text (first criteria as sample) ---")
first_crit = headers[criterias_cols[0]]
print(f"Criteria: {first_crit}")
for k, v in criterias_options[first_crit].items():
    print(f"  Option: '{k}' -> Frequency: {v}")

print("\n--- Checking if there are any options with parentheses or points ---")
all_crit_values = set()
for h in [headers[i] for i in criterias_cols]:
    for option in criterias_options[h].keys():
        all_crit_values.add(str(option))
print("Sample options from all criteria:")
for idx, opt in enumerate(sorted(list(all_crit_values))[:20]):
    print(f"  {idx}: {opt}")
