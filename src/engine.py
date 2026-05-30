import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.parser import extract_points
from src.cleaner import normalize_name, split_group_names
def map_day_to_date(day_val):
    """
    Maps day numbers (1, 2, 3, 4) to institutional defense dates starting May 26th.
    """
    if day_val is None or str(day_val).strip() == "":
        return ""
    try:
        d_num = int(float(str(day_val).strip()))
        mapping = {
            1: "Día 1 (26 de mayo)",
            2: "Día 2 (27 de mayo)",
            3: "Día 3 (28 de mayo)",
            4: "Día 4 (29 de mayo)"
        }
        return mapping.get(d_num, f"Día {d_num}")
    except ValueError:
        return f"Día {day_val}"

def is_valid_name(val):
    """
    Checks if the value is a valid name and not an empty string or None representation.
    """
    if val is None:
        return False
    val_str = str(val).strip()
    if val_str == "" or val_str.lower() in ["none", "nan", "null", "n/a", "<na>"]:
        return False
    return True

def check_replacement(evaluator_raw, submitter_name):
    """
    Checks if the person who filled out the form (submitter_name) is a replacement
    for the selected evaluator in the dropdown (evaluator_raw).
    Returns True if it's a replacement, False if it's the same person or invalid.
    """
    if not is_valid_name(evaluator_raw) or not is_valid_name(submitter_name):
        return False
    norm_eval = normalize_name(evaluator_raw)
    norm_sub = normalize_name(submitter_name)
    if norm_eval == norm_sub:
        return False
    
    # Split into words
    words_eval = set(norm_eval.split())
    words_sub = set(norm_sub.split())
    
    # If they share 2 or more words, they are likely the same person
    intersection = words_eval.intersection(words_sub)
    if len(intersection) >= 2:
        return False
        
    return True


# The 7 criteria in the exact order and with the exact names used in the historical processed file
CRITERIA_MAP = {
    "apertura": {
        "raw_pattern": "Apertura: problema y objetivo",
        "clean_name": "Promedio de Points - Apertura: problema y objetivo\xa0",
        "qual_name": "Rubrica - Apertura: problema y objetivo\xa0",
        "max_pts": 20
    },
    "metodologia": {
        "raw_pattern": "Metodología (nivel adecuado)",
        "clean_name": "Promedio de Points - Metodología (nivel adecuado)",
        "qual_name": "Rubrica -  Metodología (nivel adecuado)", # Note the double space after '-' in historical file!
        "max_pts": 20
    },
    "resultados": {
        "raw_pattern": "Resultados y evidencia",
        "clean_name": "Promedio de Points - Resultados y evidencia\xa0",
        "qual_name": "Rubrica - Resultados y evidencia\xa0",
        "max_pts": 20
    },
    "coherencia": {
        "raw_pattern": "Coherencia y manejo del tiempo",
        "clean_name": "Promedio de Points - Coherencia y manejo del tiempo",
        "qual_name": "Rubrica - Coherencia y manejo del tiempo",
        "max_pts": 10
    },
    "diapositivas": {
        "raw_pattern": "Diapositivas como apoyo",
        "clean_name": "Promedio de Points - Diapositivas como apoyo",
        "qual_name": "Rubrica - Diapositivas como apoyo",
        "max_pts": 10
    },
    "respuestas": {
        "raw_pattern": "Respuestas a preguntas",
        "clean_name": "Promedio de Points - Respuestas a preguntas",
        "qual_name": "Rubrica - Cierre: Respuestas a Preguntas", # Note the different name in historical file!
        "max_pts": 10
    },
    "cierre": {
        "raw_pattern": "Cierre: aporte a próximos pasos",
        "clean_name": "Promedio de Points - Cierre: aporte a próximos pasos",
        "qual_name": "Rubrica - Cierre: aporte a próximos pasos",
        "max_pts": 10
    }
}

def get_qualitative_label_20(score):
    """
    Returns qualitative label for 20-point criteria:
    Excelente (>=18), Muy bueno (>=14), Bueno (>=10), Regular (>=6), Insuficiente (<6)
    """
    rounded = round(score, 10)
    if rounded >= 18:
        return "Excelente"
    elif rounded >= 14:
        return "Muy bueno"
    elif rounded >= 10:
        return "Bueno"
    elif rounded >= 6:
        return "Regular"
    else:
        return "Insuficiente"

def get_qualitative_label_10(score):
    """
    Returns qualitative label for 10-point criteria:
    Excelente (>=9), Muy bueno (>=7), Bueno (>=5), Regular (>=3), Insuficiente (<3)
    """
    rounded = round(score, 10)
    if rounded >= 9:
        return "Excelente"
    elif rounded >= 7:
        return "Muy bueno"
    elif rounded >= 5:
        return "Bueno"
    elif rounded >= 3:
        return "Regular"
    else:
        return "Insuficiente"

def get_qualitative_label_final(score):
    """
    Returns qualitative label for final score (out of 100):
    Excelente (>=90), Muy Bueno (>=80), Bueno (>=75), Regular (>=65), Insuficiente (<65)
    """
    rounded = round(score, 10)
    if rounded >= 90:
        return "Excelente"
    elif rounded >= 80:
        return "Muy Bueno"  # Capital B in historical!
    elif rounded >= 75:
        return "Bueno"
    elif rounded >= 65:
        return "Regular"
    else:
        return "Insuficiente"

def find_column_by_substring(headers, substring):
    """
    Finds the column header index that contains the given substring (case insensitive).
    """
    sub = substring.lower().strip()
    for idx, h in enumerate(headers):
        if h and sub in str(h).lower():
            return idx
    return None

def process_oral_defense(raw_excel_path, schedule_excel_path=None, exclude_duplicates=False):
    """
    Processes the raw Microsoft Forms Excel sheet for Oral Defense.
    Groups evaluations by student, computes averages, maps qualitative labels,
    and returns DataFrames for individual submissions, processed calculations,
    and a committee compliance tracker.
    """
    # 1. Load Raw Excel
    wb = openpyxl.load_workbook(raw_excel_path, data_only=True)
    sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    
    # Read headers
    headers = [cell.value for cell in sheet[1]]
    
    # Find key column indices
    student_col_idx = find_column_by_substring(headers, "Seleccione el nombre del Estudiante")
    evaluator_col_idx = find_column_by_substring(headers, "Seleccione su nombre (Evaluador)")
    
    if student_col_idx is None:
        raise ValueError("No se pudo encontrar la columna 'Seleccione el nombre del Estudiante'.")
    if evaluator_col_idx is None:
        raise ValueError("No se pudo encontrar la columna 'Seleccione su nombre (Evaluador)'.")
    
    # Find criteria column indices
    criteria_indices = {}
    for key, c_info in CRITERIA_MAP.items():
        idx = find_column_by_substring(headers, c_info["raw_pattern"])
        if idx is None:
            raise ValueError(f"No se pudo encontrar la columna del criterio: '{c_info['raw_pattern']}'.")
        criteria_indices[key] = idx

    # 2. Extract Individual Records
    records = []
    for r_idx in range(2, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r_idx, column=c).value for c in range(1, len(headers) + 1)]
        # Skip if row is entirely empty
        if not any(row_vals):
            continue
            
        student_raw = row_vals[student_col_idx]
        evaluator_raw = row_vals[evaluator_col_idx]
        
        # Skip if student name is empty
        if not is_valid_name(student_raw):
            continue
            
        student_norm = normalize_name(student_raw)
        evaluator_norm = normalize_name(evaluator_raw)
        submitter_name = row_vals[4] if len(row_vals) > 4 else ""
        is_replaced = check_replacement(evaluator_raw, submitter_name)
        
        rec = {
            "Id": row_vals[0],
            "Student_Raw": student_raw,
            "Student_Normalized": student_norm,
            "Evaluator_Raw": evaluator_raw,
            "Evaluator_Normalized": evaluator_norm,
            "Email": row_vals[3] if len(row_vals) > 3 else "",
            "Submitter_Name": submitter_name if submitter_name else evaluator_raw,
            "Is_Replacement": is_replaced,
            "Date": row_vals[8] if len(row_vals) > 8 else ""
        }
        
        # Extract criteria scores
        for key, idx in criteria_indices.items():
            raw_val = row_vals[idx]
            points = extract_points(raw_val)
            rec[f"{key}_pts"] = points
            rec[f"{key}_raw"] = raw_val
            
        records.append(rec)
        
    df_individual = pd.DataFrame(records)
    
    # Optional: Exclude duplicate submissions (same evaluator + same student)
    # Keeping only the most recent one (highest Id or last row)
    if exclude_duplicates and not df_individual.empty:
        df_individual = df_individual.sort_values(by="Id", ascending=False)
        df_individual = df_individual.drop_duplicates(
            subset=["Evaluator_Normalized", "Student_Normalized"], 
            keep="first"
        )
        df_individual = df_individual.sort_values(by="Id").reset_index(drop=True)
        
    # Pre-load presentations schedule if available to apply the exactly 3 evaluators rule
    df_sched_raw = None
    if schedule_excel_path:
        try:
            sched_wb = openpyxl.load_workbook(schedule_excel_path, data_only=True)
            if "Hoja1" in sched_wb.sheetnames:
                h_sheet = sched_wb["Hoja1"]
                h_headers = [c.value for c in h_sheet[1]]
                idx_name = find_column_by_substring(h_headers, "NOMBRE")
                idx_group = find_column_by_substring(h_headers, "# GRUPO")
                idx_day = find_column_by_substring(h_headers, "DÍA DEFENSA")
                idx_hour = next((i for i, h in enumerate(h_headers) if h and str(h).strip().upper() == "HORA"), find_column_by_substring(h_headers, "HORA"))
                idx_sala = find_column_by_substring(h_headers, "SALA")
                idx_tit = find_column_by_substring(h_headers, "DOCENTE TITULACIÓN")
                idx_tutor = find_column_by_substring(h_headers, "TUTOR")
                idx_tercer = find_column_by_substring(h_headers, "TERCER DOCENTE")
                idx_adic = find_column_by_substring(h_headers, "DOCENTE ADICIONAL")
                idx_proj = find_column_by_substring(h_headers, "proyecto")
                
                students_schedule = []
                for r_idx in range(2, h_sheet.max_row + 1):
                    row_vals = [h_sheet.cell(row=r_idx, column=c).value for c in range(1, len(h_headers) + 1)]
                    if any(row_vals):
                        s_name = row_vals[idx_name] if idx_name is not None else None
                        if is_valid_name(s_name):
                            students_schedule.append({
                                "student_raw": s_name,
                                "student_norm": normalize_name(s_name),
                                "group": row_vals[idx_group] if idx_group is not None else "",
                                "day": row_vals[idx_day] if idx_day is not None else "",
                                "hour": row_vals[idx_hour] if idx_hour is not None else "",
                                "sala": row_vals[idx_sala] if idx_sala is not None else "",
                                "doc_tit": row_vals[idx_tit] if idx_tit is not None else "",
                                "doc_tutor": row_vals[idx_tutor] if idx_tutor is not None else "",
                                "doc_tercer": row_vals[idx_tercer] if idx_tercer is not None else "",
                                "doc_adic": row_vals[idx_adic] if idx_adic is not None else "",
                                "project": row_vals[idx_proj] if idx_proj is not None else ""
                            })
                df_sched_raw = pd.DataFrame(students_schedule)
        except Exception as e:
            print(f"Error pre-loading schedule in engine: {e}")

    # Enforce exactly 3 evaluations per student for Oral Defense (if N > 3)
    if not df_individual.empty and df_sched_raw is not None and not df_sched_raw.empty:
        retained_indices = []
        for s_norm, group in df_individual.groupby("Student_Normalized"):
            if len(group) <= 3:
                retained_indices.extend(group.index.tolist())
            else:
                # N > 3: We must select exactly 3 evaluations!
                # Find the planned jurors for this student from schedule
                planned_jurors = set()
                st_sched = df_sched_raw[df_sched_raw["student_norm"] == s_norm]
                if not st_sched.empty:
                    s_row = st_sched.iloc[0]
                    for f_name in ["doc_tit", "doc_tutor", "doc_tercer", "doc_adic"]:
                        if f_name in s_row and is_valid_name(s_row[f_name]):
                            planned_jurors.add(normalize_name(s_row[f_name]))
                
                # Sort group evaluations: matching planned jurors first, then earliest ID/timestamp
                group_list = []
                for idx, row in group.iterrows():
                    eval_norm = row["Evaluator_Normalized"]
                    sub_norm = normalize_name(row["Submitter_Name"]) if "Submitter_Name" in row and row["Submitter_Name"] else ""
                    
                    is_planned = False
                    if eval_norm in planned_jurors:
                        is_planned = True
                    elif sub_norm in planned_jurors:
                        is_planned = True
                        
                    group_list.append((idx, is_planned, row["Id"]))
                
                # Sort: is_planned=True first, then Id (earliest submission first)
                group_list.sort(key=lambda x: (not x[1], x[2]))
                
                # Keep exactly the top 3
                for item in group_list[:3]:
                    retained_indices.append(item[0])
                    
        df_individual = df_individual.loc[retained_indices].sort_values(by="Id").reset_index(drop=True)
    
    # 3. Consolidate by Student
    consolidated = []
    if not df_individual.empty:
        grouped = df_individual.groupby("Student_Raw") # Keep raw name as the key/display name
        
        for student_raw, group in grouped:
            student_norm = group["Student_Normalized"].iloc[0]
            eval_count = len(group)
            
            res = {
                "Seleccione el nombre del Estudiante": student_raw,
                "Cuenta de Seleccione su nombre (Evaluador)": eval_count,
            }
            
            # Compute averages and apply qualitative labels
            nota_ponderada = 0.0
            for key, c_info in CRITERIA_MAP.items():
                avg_val = group[f"{key}_pts"].mean()
                # In Excel, if the average has decimals, it remains a float.
                res[c_info["clean_name"]] = avg_val
                
                # Map qualitative labels based on score
                if c_info["max_pts"] == 20:
                    res[c_info["qual_name"]] = get_qualitative_label_20(avg_val)
                else:
                    res[c_info["qual_name"]] = get_qualitative_label_10(avg_val)
                    
                nota_ponderada += avg_val
                
            res["Nota ponderada"] = nota_ponderada
            res["Nota Rubrica Promedio"] = get_qualitative_label_final(nota_ponderada)
            
            consolidated.append(res)
        
    df_calc = pd.DataFrame(consolidated) if consolidated else pd.DataFrame(columns=[
        "Seleccione el nombre del Estudiante", "Cuenta de Seleccione su nombre (Evaluador)", "Nota ponderada", "Nota Rubrica Promedio"
    ])
    
    # Sort students alphabetically
    if not df_calc.empty:
        df_calc = df_calc.sort_values(by="Seleccione el nombre del Estudiante").reset_index(drop=True)
    
    # 4. Handle Schedule and Compliance Tracking
    compliance_records = []
    df_schedule = None
    
    # Try loading new presentations schedule format (e.g. presentaciones_crcronograma.xlsx)
    if schedule_excel_path:
        try:
            sched_wb = openpyxl.load_workbook(schedule_excel_path, data_only=True)
            if "Hoja1" in sched_wb.sheetnames:
                # This is the presentations schedule!
                h_sheet = sched_wb["Hoja1"]
                h_headers = [c.value for c in h_sheet[1]]
                
                # Find indices dynamically
                idx_name = find_column_by_substring(h_headers, "NOMBRE")
                idx_group = find_column_by_substring(h_headers, "# GRUPO")
                idx_day = find_column_by_substring(h_headers, "DÍA DEFENSA")
                idx_hour = next((i for i, h in enumerate(h_headers) if h and str(h).strip().upper() == "HORA"), find_column_by_substring(h_headers, "HORA"))
                idx_sala = find_column_by_substring(h_headers, "SALA")
                idx_tit = find_column_by_substring(h_headers, "DOCENTE TITULACIÓN")
                idx_tutor = find_column_by_substring(h_headers, "TUTOR")
                idx_tercer = find_column_by_substring(h_headers, "TERCER DOCENTE")
                idx_adic = find_column_by_substring(h_headers, "DOCENTE ADICIONAL")
                idx_proj = find_column_by_substring(h_headers, "proyecto")
                
                # Read all scheduled students
                students_schedule = []
                for r_idx in range(2, h_sheet.max_row + 1):
                    row_vals = [h_sheet.cell(row=r_idx, column=c).value for c in range(1, len(h_headers) + 1)]
                    if any(row_vals):
                        s_name = row_vals[idx_name] if idx_name is not None else None
                        if is_valid_name(s_name):
                            students_schedule.append({
                                "student_raw": s_name,
                                "student_norm": normalize_name(s_name),
                                "group": row_vals[idx_group] if idx_group is not None else "",
                                "day": row_vals[idx_day] if idx_day is not None else "",
                                "hour": row_vals[idx_hour] if idx_hour is not None else "",
                                "sala": row_vals[idx_sala] if idx_sala is not None else "",
                                "doc_tit": row_vals[idx_tit] if idx_tit is not None else "",
                                "doc_tutor": row_vals[idx_tutor] if idx_tutor is not None else "",
                                "doc_tercer": row_vals[idx_tercer] if idx_tercer is not None else "",
                                "doc_adic": row_vals[idx_adic] if idx_adic is not None else "",
                                "project": row_vals[idx_proj] if idx_proj is not None else ""
                            })
                            
                df_sched_raw = pd.DataFrame(students_schedule)
                
                if not df_sched_raw.empty:
                    # Group by group number to concatenate student names
                    grouped_sched = df_sched_raw.groupby("group")
                    sched_rows = []
                    for g_num, group_df in grouped_sched:
                        names_list = list(group_df["student_raw"].unique())
                        # Standardize with slash mapping user's cronograma sheet
                        names_str = " / ".join(names_list)
                        
                        first_row = group_df.iloc[0]
                        # We use tutor as primary docent for display
                        primary_doc = first_row["doc_tutor"] if first_row["doc_tutor"] else first_row["doc_tit"]
                        
                        sched_rows.append({
                            "Docente": primary_doc,
                            "Estudiante(s) por calificar": names_str,
                            "Día y Fecha": map_day_to_date(first_row['day']),
                            "Hora": first_row["hour"],
                            "Sala": first_row["sala"],
                            "Proyecto": first_row["project"]
                        })
                    df_schedule = pd.DataFrame(sched_rows)
                    
                    # Construct compliance records with role-based checks!
                    for _, s_row in df_sched_raw.iterrows():
                        s_norm = s_row["student_norm"]
                        s_raw = s_row["student_raw"]
                        if not is_valid_name(s_raw):
                            continue
                        day_lbl = map_day_to_date(s_row['day'])
                        
                        # Assigned docents list
                        assigned_jurors = []
                        seen_jurors_set = set()
                        for col_name, role in [
                            ("doc_tit", "Docente titulación"),
                            ("doc_tutor", "Tutor"),
                            ("doc_tercer", "Tercer docente"),
                            ("doc_adic", "Docente adicional")
                        ]:
                            juror_raw = s_row[col_name]
                            if is_valid_name(juror_raw):
                                juror_norm = normalize_name(juror_raw)
                                if juror_norm not in seen_jurors_set:
                                    seen_jurors_set.add(juror_norm)
                                    assigned_jurors.append((juror_raw, role))
                            
                        # Get all submissions for this student in df_individual
                        student_subs = df_individual[df_individual["Student_Normalized"] == s_norm].copy() if not df_individual.empty else pd.DataFrame()
                        
                        # Keep track of claimed submission indices to avoid double pairing
                        claimed_sub_indices = set()
                        
                        # Phase 1: Direct matching for planified jurors
                        juror_matches = {}
                        for juror_raw, role in assigned_jurors:
                            juror_norm = normalize_name(juror_raw)
                            matched_sub = None
                            
                            if not student_subs.empty:
                                for sub_idx, sub_row in student_subs.iterrows():
                                    if sub_idx in claimed_sub_indices:
                                        continue
                                    sub_eval_norm = sub_row["Evaluator_Normalized"]
                                    sub_subm_norm = normalize_name(sub_row["Submitter_Name"])
                                    
                                    # Direct match if the selected Evaluator dropdown matches planned juror
                                    # or the captured Submitter (M365 name) matches planned juror
                                    if sub_eval_norm == juror_norm or sub_subm_norm == juror_norm:
                                        matched_sub = (sub_idx, sub_row)
                                        break
                                        
                            if matched_sub:
                                sub_idx, sub_row = matched_sub
                                claimed_sub_indices.add(sub_idx)
                                juror_matches[juror_norm] = {
                                    "has_submitted": True,
                                    "docente_real": sub_row["Submitter_Name"],
                                    "is_replaced": sub_row["Is_Replacement"]
                                }
                                
                        # Phase 2: Smart replacement matching for planified jurors who didn't have a direct match
                        for juror_raw, role in assigned_jurors:
                            juror_norm = normalize_name(juror_raw)
                            if juror_norm in juror_matches:
                                continue # Already matched directly in Phase 1
                                
                            # Search for an unclaimed submission for this student
                            unclaimed_sub = None
                            if not student_subs.empty:
                                for sub_idx, sub_row in student_subs.iterrows():
                                    if sub_idx not in claimed_sub_indices:
                                        unclaimed_sub = (sub_idx, sub_row)
                                        break
                                        
                            if unclaimed_sub:
                                sub_idx, sub_row = unclaimed_sub
                                claimed_sub_indices.add(sub_idx)
                                # Since they don't match, this is a replacement!
                                juror_matches[juror_norm] = {
                                    "has_submitted": True,
                                    "docente_real": sub_row["Submitter_Name"],
                                    "is_replaced": True # Forced replacement
                                }
                            else:
                                juror_matches[juror_norm] = {
                                    "has_submitted": False,
                                    "docente_real": "",
                                    "is_replaced": False
                                }
                                
                        # Reconstruct group names for context
                        classmates = df_sched_raw[df_sched_raw["group"] == s_row["group"]]["student_raw"].tolist()
                        group_context = " / ".join(classmates)
                        
                        # Add to compliance records
                        for juror_raw, role in assigned_jurors:
                            juror_norm = normalize_name(juror_raw)
                            match_info = juror_matches[juror_norm]
                            
                            compliance_records.append({
                                "Docente": juror_raw,
                                "Docente_Normalized": juror_norm,
                                "Docente_Real": match_info["docente_real"] if match_info["has_submitted"] else "",
                                "Is_Replacement": match_info["is_replaced"] if match_info["has_submitted"] else False,
                                "Estudiante": s_raw,
                                "Estudiante_Normalized": s_norm,
                                "Grupo_Alumnos": group_context,
                                "Día y Fecha": day_lbl,
                                "Hora": s_row["hour"],
                                "Sala": s_row["sala"],
                                "Rol": role,
                                "Estado": "Completado" if match_info["has_submitted"] else "Pendiente",
                                "Proyecto": s_row["project"]
                            })
                            
            elif "CALIFICACION-DOCENTE" in sched_wb.sheetnames:
                # Fallback to old schedule format
                sched_sheet = sched_wb["CALIFICACION-DOCENTE"]
                sched_rows = []
                for r_idx in range(2, sched_sheet.max_row + 1):
                    row_vals = [sched_sheet.cell(row=r_idx, column=c).value for c in range(1, sched_sheet.max_column + 1)]
                    if any(row_vals):
                        doc_val = row_vals[0]
                        stud_val = row_vals[1]
                        if is_valid_name(doc_val) and is_valid_name(stud_val):
                            sched_rows.append({
                                "Docente": doc_val,
                                "Estudiante(s) por calificar": stud_val,
                                "Día y Fecha": row_vals[2],
                                "Hora": row_vals[3],
                                "Sala": row_vals[4]
                            })
                df_schedule = pd.DataFrame(sched_rows)
        except Exception as e:
            print(f"Error loading schedule: {e}")
            
    # Also fallback: if the raw_excel_path has CALIFICACION-DOCENTE sheet, load it from there!
    if df_schedule is None:
        try:
            if "CALIFICACION-DOCENTE" in wb.sheetnames:
                sched_sheet = wb["CALIFICACION-DOCENTE"]
                sched_rows = []
                for r_idx in range(2, sched_sheet.max_row + 1):
                    row_vals = [sched_sheet.cell(row=r_idx, column=c).value for c in range(1, sched_sheet.max_column + 1)]
                    if any(row_vals):
                        doc_val = row_vals[0]
                        stud_val = row_vals[1]
                        if is_valid_name(doc_val) and is_valid_name(stud_val):
                            sched_rows.append({
                                "Docente": doc_val,
                                "Estudiante(s) por calificar": stud_val,
                                "Día y Fecha": row_vals[2],
                                "Hora": row_vals[3],
                                "Sala": row_vals[4]
                            })
                df_schedule = pd.DataFrame(sched_rows)
        except Exception as e:
            print(f"Error loading schedule from raw workbook: {e}")

    # Build compliance tracker if schedule is available and no new records were populated
    if df_schedule is not None and not compliance_records:
        import re
        # Cross-reference old format
        for idx, row in df_schedule.iterrows():
            docente_raw = row["Docente"]
            students_field = row["Estudiante(s) por calificar"]
            day = row["Día y Fecha"]
            hour = row["Hora"]
            sala = row["Sala"]
            
            if is_valid_name(docente_raw) and is_valid_name(students_field):
                docente_norm = normalize_name(docente_raw)
                # Split student list
                assigned_students = split_group_names(students_field)
                
                # Check for each student if this evaluator has submitted a grade
                for s_norm in assigned_students:
                    has_submitted = False
                    docente_real = ""
                    is_replaced = False
                    if not df_individual.empty:
                        submitted_rows = df_individual[
                            (df_individual["Student_Normalized"] == s_norm) &
                            (df_individual["Evaluator_Normalized"] == docente_norm)
                        ]
                        if not submitted_rows.empty:
                            has_submitted = True
                            docente_real = submitted_rows.iloc[0]["Submitter_Name"]
                            is_replaced = submitted_rows.iloc[0]["Is_Replacement"]
                    
                    s_raw_name = ""
                    if not df_individual.empty:
                        matched_evals = df_individual[df_individual["Student_Normalized"] == s_norm]
                        if not matched_evals.empty:
                            s_raw_name = matched_evals["Student_Raw"].iloc[0]
                    
                    if not s_raw_name:
                        # Find raw name from the slash/comma separated text by searching for a match
                        chunks = [c.strip() for c in re.split(r'[/,]', str(students_field)) if c.strip()]
                        for chunk in chunks:
                            if normalize_name(chunk) == s_norm:
                                s_raw_name = chunk
                                break
                    
                    if not s_raw_name:
                        s_raw_name = s_norm # Fallback
                        
                    compliance_records.append({
                        "Docente": docente_raw,
                        "Docente_Normalized": docente_norm,
                        "Docente_Real": docente_real if has_submitted else "",
                        "Is_Replacement": is_replaced if has_submitted else False,
                        "Estudiante": s_raw_name,
                        "Estudiante_Normalized": s_norm,
                        "Grupo_Alumnos": students_field,
                        "Día y Fecha": day,
                        "Hora": hour,
                        "Sala": sala,
                        "Rol": "Jurado",
                        "Estado": "Completado" if has_submitted else "Pendiente"
                    })
                    
    df_compliance = pd.DataFrame(compliance_records) if compliance_records else pd.DataFrame(columns=[
        "Docente", "Docente_Normalized", "Docente_Real", "Is_Replacement", "Estudiante", "Estudiante_Normalized", "Grupo_Alumnos", "Día y Fecha", "Hora", "Sala", "Rol", "Estado"
    ])
    
    return df_individual, df_calc, df_compliance, df_schedule


def style_excel_sheet(ws, is_compliance=False):
    """
    Applies custom styling to an openpyxl worksheet.
    Segoe UI font, sheet-specific header colors, frozen panes, auto column widths, 
    number formats, soft-colored rows for pending, and highlights for pending teachers.
    """
    # Enable grid lines explicitly
    try:
        ws.sheet_view.showGridLines = True
    except AttributeError:
        try:
            ws.views.sheetView[0].showGridLines = True
        except Exception:
            pass
        
    sheet_title = ws.title.strip().upper() if ws.title else ""
    
    # 1. Sheet-specific Header Color (Harmonious Theme)
    # Default: Deep Navy Blue
    header_color = "1E3A8A"
    
    if "CALIFICACION GENERAL" in sheet_title:
        header_color = "0F766E"  # Muted Teal
    elif "MATRIZ" in sheet_title or "COMPLIANCE" in sheet_title:
        header_color = "374151"  # Slate Gray
    elif "CALCULO" in sheet_title or "SEGUIMIENTO" in sheet_title:
        header_color = "1E3A8A"  # Slate Navy
    elif "CALIFICACION-DOCENTE" in sheet_title:
        header_color = "4338CA"  # Elegant Indigo
    elif "PESOS" in sheet_title:
        header_color = "4B5563"  # Soft Charcoal
        
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # 1. Style Header Row
    if ws.max_row >= 1:
        ws.freeze_panes = "A2" # Freeze panes so header is static!
        ws.row_dimensions[1].height = 28
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            
    # 2. Style Data Rows
    data_font = Font(name="Segoe UI", size=10)
    data_align_left = Alignment(horizontal="left", vertical="center")
    data_align_center = Alignment(horizontal="center", vertical="center")
    
    # Status fills
    completado_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid") # Soft green
    completado_font = Font(name="Segoe UI", size=10, bold=True, color="137333") # Dark green text
    
    pendiente_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft red
    pendiente_font = Font(name="Segoe UI", size=10, bold=True, color="9C0006") # Dark red text
    
    reemplazo_fill = PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid") # Soft yellow
    reemplazo_font = Font(name="Segoe UI", size=10, bold=True, color="B06000") # Dark yellow text
    
    fila_pendiente_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid") # Very soft red for whole row
    
    # Identify key column indices dynamically for compliance sheet
    estado_col_idx = None
    docente_col_idx = None
    docente_real_col_idx = None
    if is_compliance:
        for c in range(1, ws.max_column + 1):
            header_val = str(ws.cell(row=1, column=c).value).strip().upper()
            if "ESTADO" in header_val:
                estado_col_idx = c
            elif "DOCENTE PLANIFICADO" in header_val or ("DOCENTE" in header_val and docente_col_idx is None):
                docente_col_idx = c
            elif "DOCENTE REAL" in header_val:
                docente_real_col_idx = c
                
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20
        
        # Check if the row status is pending
        is_row_pending = False
        if is_compliance and estado_col_idx is not None:
            estado_val = str(ws.cell(row=r, column=estado_col_idx).value).strip()
            if estado_val == "Pendiente":
                is_row_pending = True
                
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.border = thin_border
            
            # Text alignment and formatting
            val = cell.value
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.alignment = data_align_center
                if isinstance(val, float):
                    cell.number_format = "0.00"
            else:
                cell.alignment = data_align_left
                
            # Highlight entire row with very soft red if pending
            if is_row_pending:
                cell.fill = fila_pendiente_fill
                
            # Apply compliance status styles if compliance sheet
            if is_compliance:
                val_str = str(val).strip()
                if c == estado_col_idx:
                    if val_str == "Completado":
                        cell.fill = completado_fill
                        cell.font = completado_font
                        cell.alignment = data_align_center
                    elif val_str == "Pendiente":
                        cell.fill = pendiente_fill
                        cell.font = pendiente_font
                        cell.alignment = data_align_center
                elif c == docente_col_idx and is_row_pending:
                    # Highlight pending teacher planificado in bold red cell
                    cell.fill = pendiente_fill
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="9C0006")
                elif c == docente_real_col_idx and val_str == "Pendiente":
                    cell.fill = pendiente_fill
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="9C0006")
                elif "Sí" in val_str:
                    cell.fill = reemplazo_fill
                    cell.font = reemplazo_font
                    cell.alignment = data_align_center
                elif val_str == "No":
                    cell.alignment = data_align_center
                    
    # 3. Auto-adjust columns width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def export_to_processed_excel(df_calc, output_path, df_individual=None, df_schedule=None, df_compliance=None):
    """
    Exports the consolidated grades, qualitative rankings, and compliance matrices
    to a beautifully styled processed Excel workbook.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 1. Main Sheet1 (Raw evaluations)
        if df_individual is not None:
            df_individual.to_excel(writer, sheet_name="CALIFICACION GENERAL", index=False)
            
        # 2. Sheet Calculo
        df_calc.to_excel(writer, sheet_name="Calculo", index=False)
        df_calc.to_excel(writer, sheet_name="Seguimiento", index=False)
        
        # 3. Schedule sheet
        if df_schedule is not None:
            df_schedule.to_excel(writer, sheet_name="CALIFICACION-DOCENTE", index=False)
            
        # 4. Matriz de Seguimiento (Compliance)
        if df_compliance is not None and not df_compliance.empty:
            display_cols = ["Docente", "Rol", "Docente_Real", "Is_Replacement", "Estudiante", "Sala", "Día y Fecha", "Hora", "Estado"]
            df_comp_export = df_compliance[display_cols].copy()
            df_comp_export["¿Reemplazo?"] = df_comp_export["Is_Replacement"].apply(lambda x: "Sí ⚠️" if x else "No")
            df_comp_export["Docente Real (M365)"] = df_comp_export["Docente_Real"].apply(lambda x: x if x else "Pendiente")
            df_comp_export = df_comp_export.rename(columns={
                "Docente": "Docente Planificado",
                "Estudiante": "Estudiante",
                "Sala": "Sala",
                "Día y Fecha": "Día y Fecha",
                "Hora": "Hora",
                "Estado": "Estado"
            })
            df_comp_export = df_comp_export[["Docente Planificado", "Rol", "Docente Real (M365)", "¿Reemplazo?", "Estudiante", "Sala", "Día y Fecha", "Hora", "Estado"]]
            df_comp_export.to_excel(writer, sheet_name="Matriz de Seguimiento", index=False)
            
        # 5. PESOS Sheet
        pesos_data = [
            ("EXCELENTE", 20, 10),
            ("MUY BUENO", 16, 8),
            ("BUENO", 12, 6),
            ("REGULAR", 8, 4),
            ("INSUFICIENTE", 4, 2)
        ]
        df_pesos = pd.DataFrame(pesos_data, columns=["Nivel (A)", "Puntos_20 (B)", "Puntos_10 (C)"])
        df_pesos.to_excel(writer, sheet_name="PESOS", index=False)
        
        # Style all worksheets in the saved book
        wb = writer.book
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            is_comp = (sheetname == "Matriz de Seguimiento")
            style_excel_sheet(ws, is_compliance=is_comp)


# ==========================================
# PROYECTO CAPSTONE (INFORME ESCRITO) ENGINE
# ==========================================

WRITTEN_CRITERIA_MAP = {
    "estructura": {
        "raw_pattern": "Estructura y claridad del informe",
        "clean_name": "Promedio de Points - Estructura y claridad del informe",
        "qual_name": "Rubrica - Estructura y claridad del informe",
        "max_pts": 25
    },
    "metodologia": {
        "raw_pattern": "Metodología y coherencia",
        "clean_name": "Promedio de Points - Metodología y coherencia",
        "qual_name": "Rubrica - Metodología y coherencia",
        "max_pts": 30
    },
    "resultados": {
        "raw_pattern": "Resultados y Conclusiones",
        "clean_name": "Promedio de Points - Resultados y conclusiones",
        "qual_name": "Rubrica - Resultados y conclusiones",
        "max_pts": 25
    },
    "diapositivas": {
        "raw_pattern": "Calidad de diapositivas",
        "clean_name": "Promedio de Points - Calidad de diapositivas",
        "qual_name": "Rubrica - Calidad de diapositivas",
        "max_pts": 10
    },
    "lineamientos": {
        "raw_pattern": "Lineamientos y citación",
        "clean_name": "Promedio de Points - Lineamientos y citación",
        "qual_name": "Rubrica - Lineamientos y citación",
        "max_pts": 10
    }
}

def get_qualitative_label_percent(score, max_pts):
    """
    Returns qualitative label based on standard percentage ranges:
    Excelente (>=90%), Muy bueno (>=70%), Bueno (>=50%), Regular (>=30%), Insuficiente (<30%)
    """
    rounded = round(score, 10)
    pct = rounded / max_pts
    if pct >= 0.90:
        return "Excelente"
    elif pct >= 0.70:
        return "Muy bueno"
    elif pct >= 0.50:
        return "Bueno"
    elif pct >= 0.30:
        return "Regular"
    else:
        return "Insuficiente"

def process_capstone_written(raw_excel_path, schedule_excel_path=None, exclude_duplicates=False):
    """
    Processes the raw Microsoft Forms Excel sheet for Capstone Written Report.
    Each submission in the raw file represents a group's evaluation by a tutor.
    We split group student names, allocate the evaluation individually,
    average multiple evaluations if they exist, map qualitative labels,
    and cross-reference compliance (each scheduled student expects exactly 1 evaluation from their assigned TUTOR).
    """
    import re
    wb = openpyxl.load_workbook(raw_excel_path, data_only=True)
    sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    
    headers = [cell.value for cell in sheet[1]]
    
    # Identify key columns
    evaluator_col_idx = find_column_by_substring(headers, "Seleccione su nombre (Evaluador)")
    group_col_idx = find_column_by_substring(headers, "Seleccione el grupo a evaluar")
    project_col_idx = find_column_by_substring(headers, "Proyecto")
    
    if evaluator_col_idx is None:
        evaluator_col_idx = find_column_by_substring(headers, "Evaluador")
    if group_col_idx is None:
        group_col_idx = find_column_by_substring(headers, "grupo")
    
    if evaluator_col_idx is None or group_col_idx is None:
        raise ValueError("No se pudieron encontrar las columnas requeridas (Evaluador / Grupo).")
        
    # Criteria column indices
    criteria_indices = {}
    for key, c_info in WRITTEN_CRITERIA_MAP.items():
        idx = find_column_by_substring(headers, c_info["raw_pattern"])
        if idx is None:
            # Try a shorter substring
            short_pat = c_info["raw_pattern"].split()[0]
            idx = find_column_by_substring(headers, short_pat)
        if idx is None:
            raise ValueError(f"No se pudo encontrar la columna para el criterio: '{c_info['raw_pattern']}'.")
        criteria_indices[key] = idx
        
    # Extract records and split students
    records = []
    for r_idx in range(2, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r_idx, column=c).value for c in range(1, len(headers) + 1)]
        if not any(row_vals):
            continue
            
        evaluator_raw = row_vals[evaluator_col_idx]
        group_raw = row_vals[group_col_idx]
        project_raw = row_vals[project_col_idx] if project_col_idx is not None else ""
        
        if not is_valid_name(group_raw) or not is_valid_name(evaluator_raw):
            continue
            
        # Split names using our flexible split logic
        members_normalized = split_group_names(group_raw)
        members_raw = [m.strip() for m in re.split(r'[,/]', str(group_raw)) if m.strip()]
        
        # In case the lists differ in length, zip carefully or map raw names to normalized
        for i, s_norm in enumerate(members_normalized):
            s_raw = members_raw[i] if i < len(members_raw) else s_norm
            if not is_valid_name(s_raw):
                continue
                
            submitter_name = row_vals[4] if len(row_vals) > 4 else ""
            is_replaced = check_replacement(evaluator_raw, submitter_name)
            
            rec = {
                "Id": row_vals[0],
                "Student_Raw": s_raw,
                "Student_Normalized": s_norm,
                "Evaluator_Raw": evaluator_raw,
                "Evaluator_Normalized": normalize_name(evaluator_raw),
                "Email": row_vals[3] if len(row_vals) > 3 else "",
                "Submitter_Name": submitter_name if submitter_name else evaluator_raw,
                "Is_Replacement": is_replaced,
                "Project_Raw": project_raw,
                "Group_Raw": group_raw
            }
            
            for key, idx in criteria_indices.items():
                raw_val = row_vals[idx]
                points = extract_points(raw_val)
                rec[f"{key}_pts"] = points
                rec[f"{key}_raw"] = raw_val
                
            records.append(rec)
            
    df_individual = pd.DataFrame(records) if records else pd.DataFrame(columns=[
        "Id", "Student_Raw", "Student_Normalized", "Evaluator_Raw", "Evaluator_Normalized", "Project_Raw", "Group_Raw"
    ])
    
    # Optional: Exclude duplicate submissions (same evaluator + same student)
    # Keeping only the most recent one (highest Id or last row)
    if exclude_duplicates and not df_individual.empty:
        df_individual = df_individual.sort_values(by="Id", ascending=False)
        df_individual = df_individual.drop_duplicates(
            subset=["Evaluator_Normalized", "Student_Normalized"], 
            keep="first"
        )
        df_individual = df_individual.sort_values(by="Id").reset_index(drop=True)
    
    # Consolidate by student
    consolidated = []
    if not df_individual.empty:
        grouped = df_individual.groupby("Student_Raw")
        for student_raw, group in grouped:
            student_norm = group["Student_Normalized"].iloc[0]
            eval_count = len(group)
            
            res = {
                "Seleccione el nombre del Estudiante": student_raw,
                "Cuenta de Seleccione su nombre (Evaluador)": eval_count,
                "Proyecto": group["Project_Raw"].iloc[0] if group["Project_Raw"].iloc[0] else ""
            }
            
            nota_ponderada = 0.0
            for key, c_info in WRITTEN_CRITERIA_MAP.items():
                avg_val = group[f"{key}_pts"].mean()
                res[c_info["clean_name"]] = avg_val
                res[c_info["qual_name"]] = get_qualitative_label_percent(avg_val, c_info["max_pts"])
                nota_ponderada += avg_val
                
            res["Nota ponderada"] = nota_ponderada
            res["Nota Rubrica Promedio"] = get_qualitative_label_final(nota_ponderada)
            
            consolidated.append(res)
            
    df_calc = pd.DataFrame(consolidated) if consolidated else pd.DataFrame(columns=[
        "Seleccione el nombre del Estudiante", "Cuenta de Seleccione su nombre (Evaluador)", "Proyecto", "Nota ponderada", "Nota Rubrica Promedio"
    ])
    
    if not df_calc.empty:
        df_calc = df_calc.sort_values(by="Seleccione el nombre del Estudiante").reset_index(drop=True)
        
    # Compliance and Schedule tracking
    compliance_records = []
    df_schedule = None
    
    if schedule_excel_path:
        try:
            sched_wb = openpyxl.load_workbook(schedule_excel_path, data_only=True)
            if "Hoja1" in sched_wb.sheetnames:
                h_sheet = sched_wb["Hoja1"]
                h_headers = [c.value for c in h_sheet[1]]
                
                idx_name = find_column_by_substring(h_headers, "NOMBRE")
                idx_group = find_column_by_substring(h_headers, "# GRUPO")
                idx_day = find_column_by_substring(h_headers, "DÍA DEFENSA")
                idx_hour = next((i for i, h in enumerate(h_headers) if h and str(h).strip().upper() == "HORA"), find_column_by_substring(h_headers, "HORA"))
                idx_sala = find_column_by_substring(h_headers, "SALA")
                idx_tutor = find_column_by_substring(h_headers, "TUTOR")
                idx_proj = find_column_by_substring(h_headers, "proyecto")
                
                students_schedule = []
                for r_idx in range(2, h_sheet.max_row + 1):
                    row_vals = [h_sheet.cell(row=r_idx, column=c).value for c in range(1, len(h_headers) + 1)]
                    if any(row_vals):
                        s_name = row_vals[idx_name] if idx_name is not None else None
                        if is_valid_name(s_name):
                            students_schedule.append({
                                "student_raw": s_name,
                                "student_norm": normalize_name(s_name),
                                "group": row_vals[idx_group] if idx_group is not None else "",
                                "day": row_vals[idx_day] if idx_day is not None else "",
                                "hour": row_vals[idx_hour] if idx_hour is not None else "",
                                "sala": row_vals[idx_sala] if idx_sala is not None else "",
                                "doc_tutor": row_vals[idx_tutor] if idx_tutor is not None else "",
                                "project": row_vals[idx_proj] if idx_proj is not None else ""
                            })
                            
                df_sched_raw = pd.DataFrame(students_schedule)
                
                if not df_sched_raw.empty:
                    # Group view schedule DataFrame
                    grouped_sched = df_sched_raw.groupby("group")
                    sched_rows = []
                    for g_num, group_df in grouped_sched:
                        names_list = list(group_df["student_raw"].unique())
                        names_str = " / ".join(names_list)
                        first_row = group_df.iloc[0]
                        
                        sched_rows.append({
                            "Docente": first_row["doc_tutor"],
                            "Estudiante(s) por calificar": names_str,
                            "Día y Fecha": map_day_to_date(first_row['day']),
                            "Hora": first_row["hour"],
                            "Sala": first_row["sala"],
                            "Proyecto": first_row["project"]
                        })
                    df_schedule = pd.DataFrame(sched_rows)
                    
                    # Compliance Tracker (Tutor as the sole required evaluator!)
                    for _, s_row in df_sched_raw.iterrows():
                        s_norm = s_row["student_norm"]
                        s_raw = s_row["student_raw"]
                        if not is_valid_name(s_raw):
                            continue
                        day_lbl = map_day_to_date(s_row['day'])
                        
                        tutor_raw = s_row["doc_tutor"]
                        if is_valid_name(tutor_raw):
                            tutor_norm = normalize_name(tutor_raw)
                            has_submitted = False
                            docente_real = ""
                            is_replaced = False
                            if not df_individual.empty:
                                # First, try direct match by planned tutor name or submitter
                                submitted = df_individual[
                                    (df_individual["Student_Normalized"] == s_norm) &
                                    ((df_individual["Evaluator_Normalized"] == tutor_norm) |
                                     (df_individual["Submitter_Name"].apply(normalize_name) == tutor_norm))
                                ]
                                # Second, fallback to any unclaimed submission for this student
                                if submitted.empty:
                                    submitted = df_individual[df_individual["Student_Normalized"] == s_norm]
                                    if not submitted.empty:
                                        is_replaced = True
                                        
                                if not submitted.empty:
                                    has_submitted = True
                                    docente_real = submitted.iloc[0]["Submitter_Name"]
                                    if not is_replaced:
                                        is_replaced = submitted.iloc[0]["Is_Replacement"]
                                    
                            classmates = df_sched_raw[df_sched_raw["group"] == s_row["group"]]["student_raw"].tolist()
                            group_context = " / ".join(classmates)
                            
                            compliance_records.append({
                                "Docente": tutor_raw,
                                "Docente_Normalized": tutor_norm,
                                "Docente_Real": docente_real if has_submitted else "",
                                "Is_Replacement": is_replaced if has_submitted else False,
                                "Estudiante": s_raw,
                                "Estudiante_Normalized": s_norm,
                                "Grupo_Alumnos": group_context,
                                "Día y Fecha": day_lbl,
                                "Hora": s_row["hour"],
                                "Sala": s_row["sala"],
                                "Rol": "Tutor (Evaluador Único)",
                                "Estado": "Completado" if has_submitted else "Pendiente",
                                "Proyecto": s_row["project"]
                            })
                            
            elif "CALIFICACION-DOCENTE" in sched_wb.sheetnames:
                # Fallback to old schedule format
                sched_sheet = sched_wb["CALIFICACION-DOCENTE"]
                sched_rows = []
                for r_idx in range(2, sched_sheet.max_row + 1):
                    row_vals = [sched_sheet.cell(row=r_idx, column=c).value for c in range(1, sched_sheet.max_column + 1)]
                    if any(row_vals):
                        doc_val = row_vals[0]
                        stud_val = row_vals[1]
                        if is_valid_name(doc_val) and is_valid_name(stud_val):
                            sched_rows.append({
                                "Docente": doc_val,
                                "Estudiante(s) por calificar": stud_val,
                                "Día y Fecha": row_vals[2],
                                "Hora": row_vals[3],
                                "Sala": row_vals[4]
                            })
                df_schedule = pd.DataFrame(sched_rows)
        except Exception as e:
            print(f"Error loading schedule for written report: {e}")
            
    # Also fallback: if the raw_excel_path has CALIFICACION-DOCENTE sheet, load it from there!
    if df_schedule is None:
        try:
            if "CALIFICACION-DOCENTE" in wb.sheetnames:
                sched_sheet = wb["CALIFICACION-DOCENTE"]
                sched_rows = []
                for r_idx in range(2, sched_sheet.max_row + 1):
                    row_vals = [sched_sheet.cell(row=r_idx, column=c).value for c in range(1, sched_sheet.max_column + 1)]
                    if any(row_vals):
                        doc_val = row_vals[0]
                        stud_val = row_vals[1]
                        if is_valid_name(doc_val) and is_valid_name(stud_val):
                            sched_rows.append({
                                "Docente": doc_val,
                                "Estudiante(s) por calificar": stud_val,
                                "Día y Fecha": row_vals[2],
                                "Hora": row_vals[3],
                                "Sala": row_vals[4]
                            })
                df_schedule = pd.DataFrame(sched_rows)
        except Exception as e:
            print(f"Error loading schedule from raw workbook: {e}")

    # Build compliance tracker if schedule is available and no new records were populated
    if df_schedule is not None and not compliance_records:
        import re
        # Cross-reference old format
        for idx, row in df_schedule.iterrows():
            docente_raw = row["Docente"]
            students_field = row["Estudiante(s) por calificar"]
            day = row["Día y Fecha"]
            hour = row["Hora"]
            sala = row["Sala"]
            
            if is_valid_name(docente_raw) and is_valid_name(students_field):
                docente_norm = normalize_name(docente_raw)
                # Split student list
                assigned_students = split_group_names(students_field)
                
                # Check for each student if this evaluator has submitted a grade
                for s_norm in assigned_students:
                    has_submitted = False
                    docente_real = ""
                    is_replaced = False
                    if not df_individual.empty:
                        # First try direct match
                        submitted_rows = df_individual[
                            (df_individual["Student_Normalized"] == s_norm) &
                            ((df_individual["Evaluator_Normalized"] == docente_norm) |
                             (df_individual["Submitter_Name"].apply(normalize_name) == docente_norm))
                        ]
                        # Fallback to any submission for this student
                        if submitted_rows.empty:
                            submitted_rows = df_individual[df_individual["Student_Normalized"] == s_norm]
                            if not submitted_rows.empty:
                                is_replaced = True
                                
                        if not submitted_rows.empty:
                            has_submitted = True
                            docente_real = submitted_rows.iloc[0]["Submitter_Name"]
                            if not is_replaced:
                                is_replaced = submitted_rows.iloc[0]["Is_Replacement"]
                    
                    s_raw_name = ""
                    if not df_individual.empty:
                        matched_evals = df_individual[df_individual["Student_Normalized"] == s_norm]
                        if not matched_evals.empty:
                            s_raw_name = matched_evals["Student_Raw"].iloc[0]
                    
                    if not s_raw_name:
                        # Find raw name from the slash/comma separated text by searching for a match
                        chunks = [c.strip() for c in re.split(r'[/,]', str(students_field)) if c.strip()]
                        for chunk in chunks:
                            if normalize_name(chunk) == s_norm:
                                s_raw_name = chunk
                                break
                    
                    if not s_raw_name:
                        s_raw_name = s_norm # Fallback
                        
                    compliance_records.append({
                        "Docente": docente_raw,
                        "Docente_Normalized": docente_norm,
                        "Docente_Real": docente_real if has_submitted else "",
                        "Is_Replacement": is_replaced if has_submitted else False,
                        "Estudiante": s_raw_name,
                        "Estudiante_Normalized": s_norm,
                        "Grupo_Alumnos": students_field,
                        "Día y Fecha": day,
                        "Hora": hour,
                        "Sala": sala,
                        "Rol": "Tutor (Evaluador)",
                        "Estado": "Completado" if has_submitted else "Pendiente",
                        "Proyecto": ""
                    })
                    
    df_compliance = pd.DataFrame(compliance_records) if compliance_records else pd.DataFrame(columns=[
        "Docente", "Docente_Normalized", "Docente_Real", "Is_Replacement", "Estudiante", "Estudiante_Normalized", "Grupo_Alumnos", "Día y Fecha", "Hora", "Sala", "Rol", "Estado", "Proyecto"
    ])
    
    return df_individual, df_calc, df_compliance, df_schedule

def export_to_processed_excel_written(df_calc, output_path, df_individual=None, df_schedule=None, df_compliance=None):
    """
    Exports the consolidated Written Capstone report grades and qualitative rankings to a processed Excel workbook.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if df_individual is not None:
            df_individual.to_excel(writer, sheet_name="CALIFICACION GENERAL", index=False)
            
        df_calc.to_excel(writer, sheet_name="Calculo", index=False)
        df_calc.to_excel(writer, sheet_name="Seguimiento", index=False)
        
        if df_schedule is not None:
            df_schedule.to_excel(writer, sheet_name="CALIFICACION-DOCENTE", index=False)
            
        # 4. Matriz de Seguimiento (Compliance)
        if df_compliance is not None and not df_compliance.empty:
            display_cols = ["Docente", "Rol", "Docente_Real", "Is_Replacement", "Estudiante", "Sala", "Día y Fecha", "Hora", "Estado"]
            df_comp_export = df_compliance[display_cols].copy()
            df_comp_export["¿Reemplazo?"] = df_comp_export["Is_Replacement"].apply(lambda x: "Sí ⚠️" if x else "No")
            df_comp_export["Docente Real (M365)"] = df_comp_export["Docente_Real"].apply(lambda x: x if x else "Pendiente")
            df_comp_export = df_comp_export.rename(columns={
                "Docente": "Docente Planificado",
                "Estudiante": "Estudiante",
                "Sala": "Sala",
                "Día y Fecha": "Día y Fecha",
                "Hora": "Hora",
                "Estado": "Estado"
            })
            df_comp_export = df_comp_export[["Docente Planificado", "Rol", "Docente Real (M365)", "¿Reemplazo?", "Estudiante", "Sala", "Día y Fecha", "Hora", "Estado"]]
            df_comp_export.to_excel(writer, sheet_name="Matriz de Seguimiento", index=False)
            
        # Written Report Pesos lookup table
        pesos_data = [
            ("EXCELENTE", 25, 30, 10),
            ("MUY BUENO", 20, 24, 8),
            ("BUENO", 15, 18, 6),
            ("REGULAR", 10, 12, 4),
            ("INSUFICIENTE", 5, 6, 2)
        ]
        df_pesos = pd.DataFrame(pesos_data, columns=["Nivel (A)", "Puntos_25 (B)", "Puntos_30 (C)", "Puntos_10 (D)"])
        df_pesos.to_excel(writer, sheet_name="PESOS", index=False)
        
        # Style all worksheets in the saved book
        wb = writer.book
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            is_comp = (sheetname == "Matriz de Seguimiento")
            style_excel_sheet(ws, is_compliance=is_comp)
