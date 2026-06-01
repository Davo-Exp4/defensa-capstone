#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script de verificación y testing automatizado para la unificación grupal 
y compliance del Proyecto Capstone (Informe Escrito).
"""

import sys
import os
import openpyxl
import pandas as pd

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from src.engine import process_capstone_written, normalize_name

def run_written_verification():
    print("====================================================")
    print("EJECUTANDO VERIFICACIÓN DE INFORME ESCRITO CAPSTONE")
    print("====================================================")
    
    raw_path = "data/EVALUACIÓN PROYECTO CAPSTONE - COHORTE 2(1-59).xlsx"
    schedule_path = "data/presentaciones_crcronograma.xlsx"
    
    if not os.path.exists(raw_path) or not os.path.exists(schedule_path):
        print("[ERROR] Archivos necesarios no encontrados en data/")
        return False
        
    print("[INFO] Cargando y procesando base de datos...")
    df_ind, df_calc, df_comp, df_sched = process_capstone_written(raw_path, schedule_path)
    
    # ----------------------------------------------------
    # CHECK 1: La agregación en df_calc debe ser 100% grupal
    # ----------------------------------------------------
    print("\n--- Verificación 1: Estructura Grupal en Cálculos ---")
    print(f"Total de registros agrupados calculados: {len(df_calc)}")
    
    # Check that "Seleccione el nombre del Estudiante" contains multiple students (separated by /)
    has_group_rows = df_calc["Seleccione el nombre del Estudiante"].astype(str).str.contains("/").any()
    if has_group_rows:
        print("[OK] Las filas de cálculos se encuentran correctamente consolidadas por grupo de estudiantes.")
    else:
        print("[ERROR] No se detectaron filas consolidadas por grupo en las calificaciones calculadas.")
        return False

    # ----------------------------------------------------
    # CHECK 2: Consolidación programática de Lema y Alvear
    # ----------------------------------------------------
    print("\n--- Verificación 2: Consolidación de Lema y Alvear ---")
    
    # In df_comp (compliance matrix), search for LEMA or ALVEAR
    df_comp_matched = df_comp[df_comp["Estudiante"].str.contains("LEMA CHILIQUINGA|ALVEAR LEMA")]
    
    if df_comp_matched.empty:
        print("[ERROR] No se encontró a Lema o Alvear en la matriz de seguimiento.")
        return False
        
    print(f"Defensas de compliance encontradas para Lema/Alvear: {len(df_comp_matched)}")
    
    # Check that they occupy exactly ONE single unified row (Group 62-64)
    if len(df_comp_matched) == 1:
        row = df_comp_matched.iloc[0]
        student_val = str(row["Estudiante"])
        tutor_val = str(row["Docente"])
        proj_val = str(row["Proyecto"])
        status_val = str(row["Estado"])
        
        print(f"  - Estudiante(s) Celda: '{student_val}'")
        print(f"  - Tutor Asignado:      '{tutor_val}'")
        print(f"  - Proyecto Asignado:   '{proj_val}'")
        print(f"  - Estado de Rúbrica:   '{status_val}'")
        
        # Verify specific fields
        checks = [
            "LEMA CHILIQUINGA PAUL ALEJANDRO" in student_val,
            "ALVEAR LEMA EMILIO JOSE" in student_val,
            "SANTIAGO SOLÓRZANO LESCANO" == tutor_val,
            "Hispasat" in proj_val,
            status_val == "Completado" # Now graded in the new 59-row sheet!
        ]
        
        if all(checks):
            print("[OK] Consolidación programática de Lema y Alvear validada con éxito absoluto.")
        else:
            print("[ERROR] Mismatch en los valores consolidados programáticamente para Lema y Alvear.")
            print(f"Checks fallidos: {[i for i, c in enumerate(checks) if not c]}")
            return False
    else:
        print(f"[ERROR] Se encontraron {len(df_comp_matched)} filas de seguimiento en lugar de 1 fila unificada.")
        return False

    # ----------------------------------------------------
    # CHECK 3: Simulación de Calificación Grupal (Compliance)
    # ----------------------------------------------------
    print("\n--- Verificación 3: Simulación de Calificación para Uno ---")
    
    # Create temporary simulation
    wb = openpyxl.load_workbook(raw_path)
    sheet = wb.active
    
    # Insert a fake row at the end evaluating only one of the two (e.g. Paul Lema)
    headers = [c.value for c in sheet[1]]
    evaluator_col = next(i for i, h in enumerate(headers) if 'Evaluador' in str(h)) + 1
    group_col = next(i for i, h in enumerate(headers) if 'grupo' in str(h).lower()) + 1
    proj_col = next(i for i, h in enumerate(headers) if 'Proyecto' in str(h)) + 1
    
    new_row_idx = sheet.max_row + 1
    sheet.cell(row=new_row_idx, column=1, value=999) # Fake ID
    sheet.cell(row=new_row_idx, column=evaluator_col, value='CHRISTIAM GARZÓN')
    sheet.cell(row=new_row_idx, column=group_col, value='PONCE TAMAYO JUAN CARLOS')
    sheet.cell(row=new_row_idx, column=proj_col, value='Proyecto Ponce')
    
    # Criteria values
    for col_idx in range(10, 15):
        sheet.cell(row=new_row_idx, column=col_idx, value='EXCELENTE (30 puntos)')
        
    os.makedirs("scratch", exist_ok=True)
    temp_sim_path = "scratch/temp_verification_written.xlsx"
    wb.save(temp_sim_path)
    
    # Re-run engine with simulated file
    _, _, df_comp_sim, _ = process_capstone_written(temp_sim_path, schedule_path)
    
    row_sim = df_comp_sim[df_comp_sim["Estudiante"].str.contains("PONCE TAMAYO")].iloc[0]
    status_sim = row_sim["Estado"]
    docente_real_sim = row_sim["Docente_Real"]
    
    print(f"  - Estado Simulado:      '{status_sim}'")
    print(f"  - Docente Real:         '{docente_real_sim}'")
    
    if status_sim == "Completado" and docente_real_sim == "CHRISTIAM GARZÓN":
        print("[OK] La regla de negocio se cumple: calificar a uno de los dos valida de inmediato al grupo.")
    else:
        print("[ERROR] Falló la regla de validación cruzada. El grupo sigue pendiente o no registró al evaluador real.")
        return False
        
    # Clean up temp files
    if os.path.exists(temp_sim_path):
        os.remove(temp_sim_path)
        
    print("\n====================================================")
    print("🎉 ¡TODOS LOS SCRIPTS DE VERIFICACIÓN PASARON CON ÉXITO!")
    print("====================================================")
    return True

if __name__ == "__main__":
    success = run_written_verification()
    sys.exit(0 if success else 1)
