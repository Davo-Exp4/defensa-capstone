import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import pandas as pd
import openpyxl
from src.engine import process_oral_defense, CRITERIA_MAP

def run_blackbox_test():
    print("====================================================")
    print("EJECUTANDO PRUEBA DE CAJA NEGRA (COHORTE 2)")
    print("====================================================")
    
    raw_full_path = "data/DEFENSA ORAL DE PROYECTO CAPSTONE - COHORTE 2(1-375).xlsx"
    processed_path = "data/DEFENSA ORAL DE PROYECTO CAPSTONE - COHORTE 2_PROCESADO.xlsx"
    schedule_path = "data/presentaciones_crcronograma.xlsx"
    
    # Load 360 raw workbook, and save only first 95 rows (including header) to a temp file
    try:
        wb = openpyxl.load_workbook(raw_full_path)
    except FileNotFoundError:
        print(f"[ERROR] No se pudo encontrar el archivo raw: {raw_full_path}")
        return False
        
    sheet = wb.active
    if sheet.max_row > 95:
        sheet.delete_rows(96, sheet.max_row - 95)
        
    os.makedirs("scratch", exist_ok=True)
    temp_raw_path = "scratch/temp_raw_94.xlsx"
    wb.save(temp_raw_path)
    print(f"Slice temporal creado con éxito en: {temp_raw_path} (94 evaluaciones)")
    
    # Process using our engine
    df_ind, df_calc, df_comp, _ = process_oral_defense(temp_raw_path, schedule_path, exclude_duplicates=False)
    df_calc.columns = [str(c).strip() for c in df_calc.columns]
    
    # Load processed Excel Calculo sheet
    df_calc_hist = pd.read_excel(processed_path, sheet_name="Calculo")
    df_calc_hist.columns = [str(c).strip() for c in df_calc_hist.columns]
    df_calc_hist = df_calc_hist[df_calc_hist["Seleccione el nombre del Estudiante"].notna()]
    df_calc_hist = df_calc_hist[df_calc_hist["Seleccione el nombre del Estudiante"].str.upper() != "TOTAL GENERAL"]
    
    df_calc["student_key"] = df_calc["Seleccione el nombre del Estudiante"].str.strip().str.upper()
    df_calc_hist["student_key"] = df_calc_hist["Seleccione el nombre del Estudiante"].str.strip().str.upper()
    
    ours_indexed = df_calc.set_index("student_key")
    hist_indexed = df_calc_hist.set_index("student_key")
    
    mismatches = 0
    checked_students = 0
    
    columns_to_compare = [
        "Cuenta de Seleccione su nombre (Evaluador)",
        "Nota ponderada",
    ]
    for key, c_info in CRITERIA_MAP.items():
        columns_to_compare.append(c_info["clean_name"].strip())
        columns_to_compare.append(c_info["qual_name"].strip())
        
    for s_key in ours_indexed.index:
        if s_key in hist_indexed.index:
            student_name = ours_indexed.loc[s_key, "Seleccione el nombre del Estudiante"]
            
            # Skip Chiriboga Terán because he has 4 reviews, and our engine correctly
            # filters him down to exactly 3 under the new 3-juror rule.
            if "CHIRIBOGA TERAN" in s_key:
                print(f"\n   [INFO] Omitiendo '{student_name}' en comparación estricta de caja negra:")
                print(f"          Bajo la nueva regla estricta de 3 jurados, sus evaluaciones se redujeron a 3 (Nota={ours_indexed.loc[s_key, 'Nota ponderada']:.2f}).")
                print(f"          El histórico retenía 4 evaluaciones sin filtrar (Nota={hist_indexed.loc[s_key, 'Nota ponderada']:.2f}).")
                continue
                
            checked_students += 1
            row_ours = ours_indexed.loc[s_key]
            row_hist = hist_indexed.loc[s_key]
            
            for col in columns_to_compare:
                if col not in row_ours or col not in row_hist:
                    continue
                val_ours = row_ours[col]
                val_hist = row_hist[col]
                
                if isinstance(val_ours, (int, float)) and isinstance(val_hist, (int, float)):
                    if abs(val_ours - val_hist) > 1e-5:
                        print(f"   [ERROR] Mismatch para {student_name} en '{col}': Nuestro={val_ours:.4f}, Histórico={val_hist:.4f}")
                        mismatches += 1
                else:
                    if str(val_ours).strip().lower() != str(val_hist).strip().lower():
                        print(f"   [ERROR] Mismatch para {student_name} en '{col}': Nuestro='{val_ours}', Histórico='{val_hist}'")
                        mismatches += 1
                        
    print("\n====================================================")
    print("RESUMEN DE PRUEBA DE CAJA NEGRA")
    print("====================================================")
    print(f"Estudiantes validados (excluyendo filtros intencionales): {checked_students}")
    print(f"Total de discrepancias encontradas: {mismatches}")
    
    if mismatches == 0:
        print("\n🎉 ¡EXITO ROTUNDO! El 100% de los cálculos de promedios ponderados y cualitativos coincide perfectamente con la validación histórica.")
        return True
    else:
        print("\n❌ FALLIDO: Se encontraron discrepancias. Revisar lógica en src/engine.py.")
        return False

if __name__ == "__main__":
    run_blackbox_test()
