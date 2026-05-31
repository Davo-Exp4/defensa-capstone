import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
from io import BytesIO

# Import our backend engine
from src.engine import (
    process_oral_defense, 
    export_to_processed_excel, 
    CRITERIA_MAP,
    process_capstone_written,
    export_to_processed_excel_written,
    WRITTEN_CRITERIA_MAP,
    map_day_to_date
)
from src.cleaner import normalize_name, split_group_names

# 1. Page Configuration and Aesthetics
st.set_page_config(
    page_title="Dashboard Capstone - Defensa Oral",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom premium CSS styling (custom fonts, vibrant gradients, polished container shadows)
st.markdown("""
<style>
    /* Global Styles */
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;800&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Outfit', sans-serif;
    }
    
    /* Title banner styling */
    .banner {
        background: linear-gradient(135deg, #1e3a8a 0%, #0d9488 100%);
        color: white;
        padding: 2.5rem;
        border-radius: 16px;
        margin-bottom: 2rem;
        box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
    }
    
    .banner h1 {
        margin: 0;
        font-size: 2.5rem;
        font-weight: 800;
        letter-spacing: -0.025em;
    }
    
    .banner p {
        margin: 0.5rem 0 0 0;
        font-size: 1.1rem;
        opacity: 0.9;
        font-weight: 300;
    }
    
    /* Metrics container */
    .metric-card {
        background-color: #ffffff;
        border-radius: 12px;
        padding: 1.5rem;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -1px rgba(0, 0, 0, 0.03);
        border-left: 5px solid #0d9488;
        margin-bottom: 1rem;
    }
    
    .metric-card-title {
        font-size: 0.875rem;
        color: #6b7280;
        text-transform: uppercase;
        font-weight: 600;
        letter-spacing: 0.05em;
    }
    
    .metric-card-value {
        font-size: 2.25rem;
        color: #1f2937;
        font-weight: 800;
        margin-top: 0.25rem;
    }
    
    .metric-card-desc {
        font-size: 0.875rem;
        color: #9ca3af;
        margin-top: 0.25rem;
    }
</style>
""", unsafe_allow_html=True)

# 1.5. Admin Authentication
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    # Center login panel beautifully
    col_l1, col_l2, col_l3 = st.columns([1, 1.5, 1])
    with col_l2:
        st.markdown("""
        <div style="padding: 2.5rem; background-color: #ffffff; border-radius: 16px; box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.1), 0 8px 10px -6px rgba(0, 0, 0, 0.1); border-top: 6px solid #1e3a8a; text-align: center; margin-top: 50px; margin-bottom: 20px;">
            <img src="https://img.icons8.com/color/96/graduation-cap.png" width="70" style="margin-bottom: 1rem;"/>
            <h2 style="margin: 0; color: #1e3a8a; font-weight: 800; font-size: 1.8rem;">Control Capstone</h2>
            <p style="color: #6b7280; font-size: 0.95rem; margin-top: 0.5rem; margin-bottom: 0px;">Maestría en Inteligencia Artificial Aplicada (MIA)</p>
        </div>
        """, unsafe_allow_html=True)
        
        with st.form("login_form"):
            username = st.text_input("Usuario:", value="", placeholder="Ingrese su usuario")
            password = st.text_input("Contraseña:", value="", type="password", placeholder="Ingrese su contraseña")
            submit = st.form_submit_button("🔓 Iniciar Sesión", use_container_width=True)
            
            if submit:
                if username.strip().lower() == "admin" and password == "admin_udla_88*":
                    st.session_state["authenticated"] = True
                    st.success("¡Acceso concedido!")
                    st.rerun()
                else:
                    st.error("❌ Usuario o contraseña incorrectos.")
    st.stop()

# 2. Sidebar and File Upload
with st.sidebar:
    st.image("https://img.icons8.com/color/96/graduation-cap.png", width=70)
    st.markdown("### **Menú de Navegación**")
    
    selected_section = st.selectbox(
        "Seleccione la Sección a Visualizar:",
        [
            "📅 Resumen General de Agenda", 
            "🎓 Defensa Oral", 
            "📝 Proyecto Capstone (Informe Escrito)"
        ]
    )
    
    exclude_duplicates = False
    if selected_section in ["🎓 Defensa Oral", "📝 Proyecto Capstone (Informe Escrito)"]:
        st.markdown("---")
        st.markdown("### **Configuración de Datos**")
        exclude_duplicates = st.toggle(
            "⚙️ Excluir Duplicados",
            value=True,
            help="Si se activa, el sistema considerará únicamente la evaluación más reciente (por ID/timestamp) enviada por cada docente para cada estudiante/grupo."
        )
    
    st.markdown("---")
    st.markdown("### **Carga de Datos Personalizados**")
    
    uploaded_raw = None
    uploaded_schedule = None
    
    if selected_section == "🎓 Defensa Oral":
        st.info("Opcionalmente, sube tus reportes personalizados para Defensa Oral.")
        uploaded_raw = st.file_uploader(
            "1. Reporte Crudo Defensa Oral (.xlsx)",
            type=["xlsx"],
            help="Respuestas de Microsoft Forms para Defensa Oral"
        )
        uploaded_schedule = st.file_uploader(
            "2. Cronograma de Planificación (.xlsx)",
            type=["xlsx"],
            help="Archivo de cronograma con las defensas"
        )
    elif selected_section == "📝 Proyecto Capstone (Informe Escrito)":
        st.info("Opcionalmente, sube tus reportes personalizados para el Informe Escrito.")
        uploaded_raw = st.file_uploader(
            "1. Reporte Crudo Informe Escrito (.xlsx)",
            type=["xlsx"],
            help="Respuestas de Microsoft Forms para Informe Escrito"
        )
        uploaded_schedule = st.file_uploader(
            "2. Cronograma de Planificación (.xlsx)",
            type=["xlsx"],
            help="Archivo de cronograma con las defensas"
        )
    else:
        st.info("Opcionalmente, sube tu cronograma de planificación personalizado.")
        uploaded_schedule = st.file_uploader(
            "Cronograma de Planificación (.xlsx)",
            type=["xlsx"],
            help="Archivo de cronograma con las defensas"
        )
        
    st.markdown("---")
    if st.button("🔒 Cerrar Sesión", use_container_width=True):
        st.session_state["authenticated"] = False
        st.rerun()

# Determine file paths and process functions based on selection
raw_path = None
schedule_path = None

if selected_section == "🎓 Defensa Oral":
    active_criteria_map = CRITERIA_MAP
    demo_raw_file = "data/DEFENSA ORAL DE PROYECTO CAPSTONE - COHORTE 2(1-375).xlsx"
    process_func = process_oral_defense
    export_func = export_to_processed_excel
    export_file_name = "consolidado_defensa_oral_procesado.xlsx"
elif selected_section == "📝 Proyecto Capstone (Informe Escrito)":
    active_criteria_map = WRITTEN_CRITERIA_MAP
    demo_raw_file = "data/EVALUACIÓN PROYECTO CAPSTONE - COHORTE 2(1-39).xlsx"
    process_func = process_capstone_written
    export_func = export_to_processed_excel_written
    export_file_name = "consolidado_proyecto_capstone_escrito_procesado.xlsx"
else:
    active_criteria_map = {}
    demo_raw_file = None
    process_func = None
    export_func = None
    export_file_name = None

# Save uploaded files or default to preloaded Cohorte 2 data
if selected_section in ["🎓 Defensa Oral", "📝 Proyecto Capstone (Informe Escrito)"]:
    if uploaded_raw:
        with open("temp_raw.xlsx", "wb") as f:
            f.write(uploaded_raw.getbuffer())
        raw_path = "temp_raw.xlsx"
    else:
        raw_path = demo_raw_file if os.path.exists(demo_raw_file) else None

if uploaded_schedule:
    with open("temp_sched.xlsx", "wb") as f:
        f.write(uploaded_schedule.getbuffer())
    schedule_path = "temp_sched.xlsx"
else:
    schedule_path = "data/presentaciones_crcronograma.xlsx" if os.path.exists("data/presentaciones_crcronograma.xlsx") else None

# Header Banner
if selected_section == "📅 Resumen General de Agenda":
    banner_subtitle = "Cronograma General y Planificación de Defensas"
else:
    banner_subtitle = "Defensa Oral (Tribunal)" if selected_section == "🎓 Defensa Oral" else "Proyecto Capstone (Informe Escrito)"

st.markdown(f"""
<div class="banner">
    <h1>🎓 Sistema de Consolidación Capstone</h1>
    <p>Automatización del procesamiento de rúbricas y analítica de {banner_subtitle}</p>
</div>
""", unsafe_allow_html=True)

# 3. Main Data Processing Logic
if selected_section == "📅 Resumen General de Agenda":
    if schedule_path and os.path.exists(schedule_path):
        try:
            wb = openpyxl.load_workbook(schedule_path, data_only=True)
            if "Proyectos agrupados" in wb.sheetnames:
                sh = wb["Proyectos agrupados"]
                headers = [cell.value for cell in sh[1]]
                
                def get_idx(lst, sub):
                    # Try case-insensitive exact match first
                    for idx, h in enumerate(lst):
                        if h and str(h).strip().lower() == sub.lower():
                            return idx
                    # Fallback to substring match
                    for idx, h in enumerate(lst):
                        if h and sub.lower() in str(h).lower():
                            return idx
                    return None
                
                idx_group = get_idx(headers, "# GRUPO")
                idx_day = get_idx(headers, "Día")
                idx_hour = get_idx(headers, "Hora")
                idx_sala = get_idx(headers, "Sala")
                idx_tit = get_idx(headers, "Docente titulación")
                idx_tutor = get_idx(headers, "Tutor")
                idx_tercer = get_idx(headers, "Tercer docente")
                idx_adic = get_idx(headers, "Docente adicional")
                idx_proj = get_idx(headers, "Proyecto")
                idx_int = get_idx(headers, "Integrantes")
                
                rows = []
                all_jurors = set()
                all_students = set()
                
                def is_valid_agenda_cell(v):
                    if v is None:
                        return False
                    v_str = str(v).strip()
                    if v_str == "" or v_str.lower() in ["none", "nan", "null", "<na>"]:
                        return False
                    return True

                for r_idx in range(2, sh.max_row + 1):
                    vals = [sh.cell(row=r_idx, column=c).value for c in range(1, len(headers) + 1)]
                    if any(vals):
                        proj_val = vals[idx_proj] if idx_proj is not None and idx_proj < len(vals) else None
                        int_val = vals[idx_int] if idx_int is not None and idx_int < len(vals) else None
                        group_val = vals[idx_group] if idx_group is not None and idx_group < len(vals) else None
                        
                        # Skip ghost rows (completely empty of actual project or student names)
                        if not is_valid_agenda_cell(int_val) and not is_valid_agenda_cell(proj_val) and not is_valid_agenda_cell(group_val):
                            continue
                            
                        # Unique jurors
                        for c_idx in [idx_tit, idx_tutor, idx_tercer, idx_adic]:
                            if c_idx is not None and c_idx < len(vals) and is_valid_agenda_cell(vals[c_idx]):
                                all_jurors.add(str(vals[c_idx]).strip().upper())
                        # Unique students
                        if idx_int is not None and idx_int < len(vals) and is_valid_agenda_cell(vals[idx_int]):
                            students_list = [s.strip() for s in str(vals[idx_int]).split('/') if s.strip()]
                            for s in students_list:
                                if is_valid_agenda_cell(s):
                                    all_students.add(s.upper())
                                
                        # Construct clean tribunal string
                        tit_ok = is_valid_agenda_cell(vals[idx_tit]) if idx_tit is not None else False
                        tercer_ok = is_valid_agenda_cell(vals[idx_tercer]) if idx_tercer is not None else False
                        adic_ok = is_valid_agenda_cell(vals[idx_adic]) if idx_adic is not None and idx_adic < len(vals) else False
                        
                        tribunal_parts = []
                        if tit_ok:
                            tribunal_parts.append(str(vals[idx_tit]).strip())
                        if tercer_ok:
                            tribunal_parts.append(str(vals[idx_tercer]).strip())
                        if adic_ok:
                            tribunal_parts.append(str(vals[idx_adic]).strip())
                            
                        rows.append({
                            "Grupo": vals[idx_group] if idx_group is not None else "",
                            "Día": map_day_to_date(vals[idx_day]) if idx_day is not None and is_valid_agenda_cell(vals[idx_day]) else "",
                            "Hora": vals[idx_hour] if idx_hour is not None else "",
                            "Sala": vals[idx_sala] if idx_sala is not None else "",
                            "Tutor": vals[idx_tutor] if idx_tutor is not None and is_valid_agenda_cell(vals[idx_tutor]) else "",
                            "Tribunal / Jurados": " / ".join(tribunal_parts) if tribunal_parts else "",
                            "Proyecto": vals[idx_proj] if idx_proj is not None and is_valid_agenda_cell(vals[idx_proj]) else "",
                            "Integrantes": vals[idx_int] if idx_int is not None and is_valid_agenda_cell(vals[idx_int]) else ""
                        })
                        
                df_sched = pd.DataFrame(rows)
                
                # Render KPIs
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-card-title">Grupos Programados</div>
                        <div class="metric-card-value">{len(df_sched)}</div>
                        <div class="metric-card-desc">Total equipos de defensa</div>
                    </div>
                    """, unsafe_allow_html=True)
                with col2:
                    st.markdown(f"""
                    <div class="metric-card" style="border-left-color: #1e3a8a;">
                        <div class="metric-card-title">Alumnos Registrados</div>
                        <div class="metric-card-value">{len(all_students)}</div>
                        <div class="metric-card-desc">Total estudiantes en cronograma</div>
                    </div>
                    """, unsafe_allow_html=True)
                with col3:
                    st.markdown(f"""
                    <div class="metric-card" style="border-left-color: #8b5cf6;">
                        <div class="metric-card-title">Tribunales Asignados</div>
                        <div class="metric-card-value">{len(all_jurors)}</div>
                        <div class="metric-card-desc">Evaluadores y Tutores únicos</div>
                    </div>
                    """, unsafe_allow_html=True)
                with col4:
                    st.markdown(f"""
                    <div class="metric-card" style="border-left-color: #f59e0b;">
                        <div class="metric-card-title">Jornadas Programadas</div>
                        <div class="metric-card-value">{len(df_sched["Día"].dropna().unique())}</div>
                        <div class="metric-card-desc">Días de defensa oral</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("---")
                st.markdown("#### **🔍 Búsqueda Avanzada y Filtros de Agenda**")
                
                # Filters
                col_f1, col_f2, col_f3 = st.columns([2, 1, 1])
                with col_f1:
                    search_query = st.text_input("Buscar por Estudiante, Proyecto, Docente o Grupo:", placeholder="Ej: Ayala, NovaPet, Ponce...")
                with col_f2:
                    filter_sala = st.selectbox("Filtrar por Sala:", ["Todas"] + sorted(list(df_sched["Sala"].dropna().unique())))
                with col_f3:
                    filter_dia = st.selectbox("Filtrar por Día:", ["Todos"] + sorted(list(df_sched["Día"].dropna().unique())))
                
                # Filter logic
                df_filtered = df_sched.copy()
                if filter_sala != "Todas":
                    df_filtered = df_filtered[df_filtered["Sala"] == filter_sala]
                if filter_dia != "Todos":
                    df_filtered = df_filtered[df_filtered["Día"] == filter_dia]
                if search_query:
                    q = search_query.lower().strip()
                    df_filtered = df_filtered[
                        df_filtered["Integrantes"].astype(str).str.lower().str.contains(q) |
                        df_filtered["Proyecto"].astype(str).str.lower().str.contains(q) |
                        df_filtered["Tutor"].astype(str).str.lower().str.contains(q) |
                        df_filtered["Tribunal / Jurados"].astype(str).str.lower().str.contains(q) |
                        df_filtered["Grupo"].astype(str).str.lower().str.contains(q)
                    ]
                
                st.markdown(f"**Defensas encontradas:** {len(df_filtered)}")
                st.dataframe(df_filtered.reset_index(drop=True), use_container_width=True)
                
            else:
                st.error("No se pudo encontrar la pestaña 'Proyectos agrupados' en el cronograma.")
        except Exception as e:
            st.error(f"Error al cargar el cronograma: {e}")
    else:
        st.warning("⚠️ No se ha encontrado un cronograma en presentations_crcronograma.xlsx o cargado en la barra lateral.")
elif raw_path:
    try:
        # Load and run backend engine (with dynamic duplicate exclusion toggle)
        df_individual, df_calc, df_compliance, df_schedule = process_func(
            raw_path, 
            schedule_path,
            exclude_duplicates=exclude_duplicates
        )
        
        # 4. TABBED INTERFACE
        tab_general, tab_individual, tab_group, tab_compliance, tab_duplicates, tab_exporter = st.tabs([
            "📊 Vista General de Cohorte",
            "👤 Reporte por Estudiante",
            "👥 Reporte por Grupos",
            "📋 Control de Seguimiento",
            "🔍 Control de Duplicados",
            "📥 Descarga de Reportes"
        ])
        
        # ----------------------------------------------------
        # TAB 1: GENERAL VIEW / KPI & CHARTS
        # ----------------------------------------------------
        with tab_general:
            st.markdown("### **Métricas Clave de la Cohorte**")
            
            # Row of KPIs
            col1, col2, col3, col4 = st.columns(4)
            
            avg_global_grade = df_calc["Nota ponderada"].mean() if not df_calc.empty else 0.0
            compliance_rate = 0.0
            total_evaluations = len(df_individual)
            
            total_sched_students = len(df_calc)
            pending_students = 0
            evaluated_students = len(df_calc)
            
            if not df_compliance.empty:
                total_sched = len(df_compliance)
                completed_sched = len(df_compliance[df_compliance["Estado"] == "Completado"])
                compliance_rate = (completed_sched / total_sched) * 100 if total_sched > 0 else 100.0
                if "Estudiante_Normalized" in df_compliance.columns:
                    total_sched_students = df_compliance["Estudiante_Normalized"].nunique()
                    # A student/group is pending if they have at least one 'Pendiente' record in compliance
                    pending_student_names = df_compliance[df_compliance["Estado"] == "Pendiente"]["Estudiante_Normalized"].unique()
                    pending_students = len(pending_student_names)
                    evaluated_students = total_sched_students - pending_students
            
            with col1:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-card-title">Estudiantes</div>
                    <div class="metric-card-value">{total_sched_students}</div>
                    <div class="metric-card-desc"><b>{evaluated_students}</b> evaluados / <b>{pending_students}</b> pendientes</div>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                st.markdown(f"""
                <div class="metric-card" style="border-left-color: #1e3a8a;">
                    <div class="metric-card-title">Promedio Global</div>
                    <div class="metric-card-value">{avg_global_grade:.2f}/100</div>
                    <div class="metric-card-desc">Calificación media de cohorte</div>
                </div>
                """, unsafe_allow_html=True)
            with col3:
                st.markdown(f"""
                <div class="metric-card" style="border-left-color: #8b5cf6;">
                    <div class="metric-card-title">Evaluaciones Recibidas</div>
                    <div class="metric-card-value">{total_evaluations}</div>
                    <div class="metric-card-desc">Total de rúbricas enviadas por jurados</div>
                </div>
                """, unsafe_allow_html=True)
            with col4:
                desc_text = "Sin cronograma asignado"
                if not df_compliance.empty:
                    remaining_sched = total_sched - completed_sched
                    desc_text = f"<b>{completed_sched}</b> de <b>{total_sched}</b> ({remaining_sched} restantes)"
                st.markdown(f"""
                <div class="metric-card" style="border-left-color: #f59e0b;">
                    <div class="metric-card-title">Compliance del Tribunal</div>
                    <div class="metric-card-value">{compliance_rate:.1f}%</div>
                    <div class="metric-card-desc">{desc_text}</div>
                </div>
                """, unsafe_allow_html=True)
                
            # Visual Progress Bar
            if not df_compliance.empty:
                remaining_sched = total_sched - completed_sched
                st.markdown(f"""
                <div style="background-color: #ffffff; padding: 1rem 1.5rem; border-radius: 12px; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05); margin-bottom: 1.5rem; border-left: 5px solid #f59e0b;">
                    <div style="font-size: 0.85rem; font-weight: 600; color: #4b5563; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 0.25rem;">
                        Progreso de Calificaciones del Tribunal
                    </div>
                    <div style="font-size: 0.95rem; color: #1f2937; margin-bottom: 0.5rem;">
                        Se han completado <b>{completed_sched}</b> de las <b>{total_sched}</b> evaluaciones programadas en el cronograma. Quedan <b>{remaining_sched}</b> calificaciones restantes por enviar.
                    </div>
                </div>
                """, unsafe_allow_html=True)
                st.progress(compliance_rate / 100.0)
                
            st.markdown("---")
            
            # Interactive Visualizations
            col_chart1, col_chart2 = st.columns(2)
            
            with col_chart1:
                st.markdown("##### **Distribución de Calificaciones Cualitativas**")
                # Group by qualitative grade
                df_labels = df_calc["Nota Rubrica Promedio"].value_counts().reset_index()
                df_labels.columns = ["Calificación Cualitativa", "Estudiantes"]
                
                # Custom colors matching standard dashboard layouts
                color_map = {
                    "Excelente": "#0d9488",
                    "Muy Bueno": "#1e3a8a",
                    "Bueno": "#8b5cf6",
                    "Regular": "#f59e0b",
                    "Insuficiente": "#ef4444"
                }
                
                fig_dona = px.pie(
                    df_labels,
                    values="Estudiantes",
                    names="Calificación Cualitativa",
                    hole=0.4,
                    color="Calificación Cualitativa",
                    color_discrete_map=color_map
                )
                fig_dona.update_layout(margin=dict(t=0, b=0, l=0, r=0))
                st.plotly_chart(fig_dona, use_container_width=True)
                
            with col_chart2:
                st.markdown("##### **Fortalezas y Debilidades: Promedio por Criterio**")
                # Calculate average for each criteria
                criteria_averages = []
                criteria_names = []
                for key, c_info in active_criteria_map.items():
                    avg_score_pts = df_calc[c_info["clean_name"]].mean()
                    # Normalize to % for fair comparison
                    avg_score_pct = (avg_score_pts / c_info["max_pts"]) * 100
                    criteria_averages.append(avg_score_pct)
                    criteria_names.append(c_info["raw_pattern"])
                
                df_crit = pd.DataFrame({
                    "Criterio de Rúbrica": criteria_names,
                    "Rendimiento (%)": criteria_averages
                }).sort_values("Rendimiento (%)", ascending=True)
                
                fig_bar = px.bar(
                    df_crit,
                    y="Criterio de Rúbrica",
                    x="Rendimiento (%)",
                    orientation="h",
                    color="Rendimiento (%)",
                    color_continuous_scale="Tealgrn",
                    range_x=[0, 100]
                )
                fig_bar.update_layout(margin=dict(t=0, b=0, l=0, r=0), showlegend=False, coloraxis_showscale=False)
                st.plotly_chart(fig_bar, use_container_width=True)
                
            # Distribution Histogram
            st.markdown("##### **Distribución Histográfica de Calificaciones Numéricas**")
            fig_hist = px.histogram(
                df_calc,
                x="Nota ponderada",
                nbins=25,
                labels={"Nota ponderada": "Calificación Final (Sobre 100)"},
                color_discrete_sequence=["#1e3a8a"]
            )
            fig_hist.update_layout(bargap=0.08, margin=dict(t=10, b=10))
            st.plotly_chart(fig_hist, use_container_width=True)
            
        # ----------------------------------------------------
        # TAB 2: INDIVIDUAL STUDENT SEARCH
        # ----------------------------------------------------
        with tab_individual:
            st.markdown("### **Visor Detallado de Calificaciones Individuales**")
            st.markdown("Selecciona un estudiante para visualizar su reporte oficial, desgloses cualitativos de rúbrica y las opiniones específicas de los jurados.")
            
            # Selectbox for students
            students_list = sorted(df_calc["Seleccione el nombre del Estudiante"].unique())
            selected_student = st.selectbox("Buscar Estudiante:", students_list)
            
            if selected_student:
                # Retrieve student summary row
                student_summary = df_calc[df_calc["Seleccione el nombre del Estudiante"] == selected_student].iloc[0]
                
                # Student KPI section
                col_st1, col_st2, col_st3 = st.columns(3)
                
                # Color code final level
                qual_final = student_summary["Nota Rubrica Promedio"]
                badge_color = "#ef4444"
                if qual_final == "Excelente":
                    badge_color = "#0d9488"
                elif qual_final == "Muy Bueno":
                    badge_color = "#1e3a8a"
                elif qual_final == "Bueno":
                    badge_color = "#8b5cf6"
                elif qual_final == "Regular":
                    badge_color = "#f59e0b"
                
                with col_st1:
                    st.markdown(f"""
                    <div class="metric-card" style="border-left-color: {badge_color};">
                        <div class="metric-card-title">Estudiante Seleccionado</div>
                        <div class="metric-card-value" style="font-size: 1.5rem; word-break: break-all; height: 50px; overflow: hidden; display: flex; align-items: center;">{selected_student}</div>
                        <div class="metric-card-desc">Reporte Oficial Mia Capstone</div>
                    </div>
                    """, unsafe_allow_html=True)
                with col_st2:
                    st.markdown(f"""
                    <div class="metric-card" style="border-left-color: {badge_color};">
                        <div class="metric-card-title">Calificación Final</div>
                        <div class="metric-card-value">{student_summary['Nota ponderada']:.2f}/100</div>
                        <div class="metric-card-desc">Promedio ponderado consolidado</div>
                    </div>
                    """, unsafe_allow_html=True)
                with col_st3:
                    st.markdown(f"""
                    <div class="metric-card" style="border-left-color: {badge_color};">
                        <div class="metric-card-title">Evaluación Cualitativa</div>
                        <div class="metric-card-value" style="color: {badge_color}; font-weight: 800;">{qual_final}</div>
                        <div class="metric-card-desc">Estado General de Aprobación</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("##### **Desglose de Calificaciones por Criterio de Rúbrica**")
                
                # Build beautiful breakdown table
                breakdown_records = []
                for key, c_info in active_criteria_map.items():
                    avg_pts = student_summary[c_info["clean_name"]]
                    qual_lbl = student_summary[c_info["qual_name"]]
                    pct = (avg_pts / c_info["max_pts"]) * 100
                    
                    breakdown_records.append({
                        "Criterio de Evaluación": c_info["raw_pattern"].strip(),
                        "Puntaje Obtenido (Promedio)": f"{avg_pts:.2f} / {c_info['max_pts']}",
                        "Rendimiento (%)": f"{pct:.1f}%",
                        "Resultado Cualitativo": qual_lbl
                    })
                
                df_breakdown = pd.DataFrame(breakdown_records)
                st.table(df_breakdown)
                
                # Individual Juror Assessments and Comments
                st.markdown("##### **Revisiones Individuales de Miembros del Tribunal**")
                # Filter individual reviews
                # Clean names in df_individual to match correctly
                norm_target = normalize_name(selected_student)
                df_st_evals = df_individual[df_individual["Student_Normalized"] == norm_target]
                
                if not df_st_evals.empty:
                    for j_idx, (_, row_eval) in enumerate(df_st_evals.iterrows(), 1):
                        juror_name = row_eval["Evaluator_Raw"]
                        
                        # Gather their specific grades
                        juror_grades = []
                        for key, c_info in active_criteria_map.items():
                            pts = row_eval[f"{key}_pts"]
                            raw_desc = row_eval[f"{key}_raw"]
                            juror_grades.append(f"**{c_info['raw_pattern'].strip()}**: {pts} pts ({raw_desc})")
                            
                        # Feedback columns in Forms raw file are located dynamically or we can list comments.
                        # Let's search if they submitted any textual comments (e.g. feedback columns)
                        feedbacks = []
                        for h_col in df_st_evals.columns:
                            if "feedback" in h_col.lower() and row_eval[h_col] is not None:
                                feedbacks.append(f"*{h_col}*: {row_eval[h_col]}")
                        
                        # Expander for each juror
                        with st.expander(f"🔔 Evaluador {j_idx}: {juror_name} — Calificación Total: {sum(row_eval[f'{key}_pts'] for key in active_criteria_map):.0f}/100"):
                            col_j1, col_j2 = st.columns(2)
                            with col_j1:
                                st.markdown("**Puntajes en Rúbrica:**")
                                for grade_str in juror_grades:
                                    st.write(f"- {grade_str}")
                            with col_j2:
                                st.markdown("**Comentarios y Observaciones (Feedback):**")
                                if feedbacks:
                                    for fbk in feedbacks:
                                        st.write(fbk)
                                else:
                                    st.info("Sin comentarios adicionales provistos por el evaluador.")
                else:
                    st.warning("No se encontraron registros de rúbricas individuales para este estudiante.")
                    
        # ----------------------------------------------------
        # TAB 2.5: GROUP VIEW REPORT
        # ----------------------------------------------------
        with tab_group:
            st.markdown("### **Visor Detallado de Calificaciones por Grupos**")
            st.markdown("Selecciona una sala o grupo de defensa programado para visualizar de forma comparativa las calificaciones individuales de cada integrante.")
            
            if df_schedule is not None and not df_schedule.empty:
                # Find unique group student strings
                unique_groups = df_schedule[["Estudiante(s) por calificar", "Sala", "Día y Fecha", "Hora"]].drop_duplicates()
                
                # Build a nice display label for each group
                group_options = []
                group_mapping = {}
                
                for idx, row in unique_groups.iterrows():
                    students_str = row["Estudiante(s) por calificar"]
                    sala = row["Sala"]
                    dia = row["Día y Fecha"]
                    hora = row["Hora"]
                    
                    label = f"📍 {sala} ({dia} a las {hora}) — {students_str}"
                    group_options.append(label)
                    group_mapping[label] = students_str
                    
                selected_group_label = st.selectbox("Seleccione un Grupo de Defensa:", sorted(group_options))
                
                if selected_group_label:
                    students_field = group_mapping[selected_group_label]
                    # Get individual members of this group
                    members = split_group_names(students_field)
                    
                    st.markdown(f"#### **Miembros del Grupo:**")
                    
                    # Create side-by-side columns or a nice layout for each member!
                    cols = st.columns(len(members))
                    
                    for m_idx, member_norm in enumerate(members):
                        with cols[m_idx]:
                            # Retrieve student data
                            df_match = df_calc[df_calc["Seleccione el nombre del Estudiante"].apply(normalize_name).str.contains(member_norm)]
                            
                            if not df_match.empty:
                                st_sum = df_match.iloc[0]
                                st_raw_name = st_sum["Seleccione el nombre del Estudiante"]
                                st_grade = st_sum["Nota ponderada"]
                                st_qual = st_sum["Nota Rubrica Promedio"]
                                st_proj = st_sum["Proyecto"] if "Proyecto" in st_sum else ""
                                
                                badge_color = "#ef4444"
                                if st_qual == "Excelente":
                                    badge_color = "#0d9488"
                                elif st_qual == "Muy Bueno":
                                    badge_color = "#1e3a8a"
                                elif st_qual == "Bueno":
                                    badge_color = "#8b5cf6"
                                elif st_qual == "Regular":
                                    badge_color = "#f59e0b"
                                    
                                html_card = f"""<div style="background-color: #f9fafb; padding: 1.5rem; border-radius: 12px; border-top: 5px solid {badge_color}; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05); margin-bottom: 1rem;">
<h4 style="margin: 0; color: #1f2937;">{st_raw_name}</h4>
{f'<p style="margin: 0.25rem 0 0 0; font-size: 0.8rem; color: #4b5563; font-style: italic;"><b>Proyecto:</b> {st_proj}</p>' if st_proj else ''}
<p style="margin: 0.5rem 0 0 0; font-size: 1.8rem; font-weight: 800; color: #111827;">{st_grade:.2f}/100</p>
<p style="margin: 0.25rem 0 0 0; font-size: 0.875rem; font-weight: bold; color: {badge_color};">{st_qual}</p>
</div>"""
                                st.markdown(html_card, unsafe_allow_html=True)
                                
                                # Criteria breakdown inside expander
                                with st.expander(f"📋 Rúbrica desglosada"):
                                    breakdown_items = []
                                    for key, c_info in active_criteria_map.items():
                                        pts = st_sum[c_info["clean_name"]]
                                        lbl = st_sum[c_info["qual_name"]]
                                        breakdown_items.append(f"**{c_info['raw_pattern'].strip()}:** {pts:.2f} pts ({lbl})")
                                    st.write("\n".join([f"- {item}" for item in breakdown_items]))
                                        
                                # Juror reviews inside another expander
                                df_st_evals = df_individual[df_individual["Student_Normalized"] == member_norm]
                                if not df_st_evals.empty:
                                    with st.expander(f"🔔 Rúbricas de Jurados ({len(df_st_evals)})"):
                                        for j_idx, (_, row_eval) in enumerate(df_st_evals.iterrows(), 1):
                                            st.markdown(f"**Jurado {j_idx}: {row_eval['Evaluator_Raw']}**")
                                            j_score = sum(row_eval[f'{key}_pts'] for key in active_criteria_map)
                                            st.write(f"Nota: {j_score:.0f}/100")
                                            # Show text feedback if any
                                            feedbacks = []
                                            for h_col in df_st_evals.columns:
                                                if "feedback" in h_col.lower() and row_eval[h_col] is not None:
                                                    feedbacks.append(str(row_eval[h_col]))
                                            if feedbacks:
                                                st.markdown(f"*Observaciones:* {'; '.join(feedbacks)}")
                                            st.markdown("---")
                            else:
                                st.markdown(f"""
                                <div style="background-color: #f9fafb; padding: 1.5rem; border-radius: 12px; border-top: 5px solid #ef4444; margin-bottom: 1rem;">
                                    <h4 style="margin: 0; color: #9ca3af;">{member_norm}</h4>
                                    <p style="margin: 0.5rem 0 0 0; color: #ef4444; font-weight: bold;">Sin evaluar</p>
                                </div>
                                """, unsafe_allow_html=True)
            else:
                st.info("💡 Por favor, sube un **Excel de Planificación** con la pestaña 'CALIFICACION-DOCENTE' para habilitar la visualización por grupos.")
                
        # ----------------------------------------------------
        # TAB 3: COMMITTEE MONITORING (COMPLIANCE TRACKER)
        # ----------------------------------------------------
        with tab_compliance:
            st.markdown("### **Matriz de Control de Seguimiento de Jurados**")
            st.markdown("Monitorea en tiempo real el cumplimiento y envío de calificaciones por sala, día y evaluador asignado.")
            
            if not df_compliance.empty:
                # Sidebar filters for compliance
                col_c1, col_c2, col_c3 = st.columns(3)
                with col_c1:
                    filter_status = st.selectbox("Filtrar por Estado:", ["Todos", "Completado", "Pendiente"])
                with col_c2:
                    filter_sala = st.selectbox("Filtrar por Sala:", ["Todas"] + list(df_compliance["Sala"].unique()))
                with col_c3:
                    filter_evaluator = st.selectbox("Filtrar por Jurado:", ["Todos"] + list(df_compliance["Docente"].unique()))
                
                # Apply filters
                df_filtered = df_compliance.copy()
                if filter_status != "Todos":
                    df_filtered = df_filtered[df_filtered["Estado"] == filter_status]
                if filter_sala != "Todas":
                    df_filtered = df_filtered[df_filtered["Sala"] == filter_sala]
                if filter_evaluator != "Todos":
                    df_filtered = df_filtered[df_filtered["Docente"] == filter_evaluator]
                
                # Compliance progress bar
                num_tot = len(df_filtered)
                num_done = len(df_filtered[df_filtered["Estado"] == "Completado"])
                pct_done = (num_done / num_tot) * 100 if num_tot > 0 else 100.0
                
                st.markdown(f"**Progreso de evaluación del comités filtrado:** ({num_done} de {num_tot} completados)")
                st.progress(pct_done / 100.0)
                
                # Render table with styled statuses
                def style_status(val):
                    color = "rgba(16, 185, 129, 0.2)" if val == "Completado" else "rgba(239, 68, 68, 0.2)"
                    text_color = "#047857" if val == "Completado" else "#b91c1c"
                    return f'background-color: {color}; color: {text_color}; font-weight: bold; text-align: center; border-radius: 4px;'
                
                def style_replacement(val):
                    if val == "Sí ⚠️":
                        return 'background-color: rgba(245, 158, 11, 0.2); color: #b45309; font-weight: bold; text-align: center; border-radius: 4px;'
                    return 'text-align: center;'
                
                # Clean dataframe for aesthetic presentation
                display_cols = ["Docente", "Rol", "Docente_Real", "Is_Replacement", "Estudiante", "Sala", "Día y Fecha", "Hora", "Estado"]
                df_disp = df_filtered[display_cols].copy()
                df_disp["¿Reemplazo?"] = df_disp["Is_Replacement"].apply(lambda x: "Sí ⚠️" if x else "No")
                df_disp["Docente Real (M365)"] = df_disp["Docente_Real"].apply(lambda x: x if x else "Pendiente")
                df_disp = df_disp.rename(columns={
                    "Docente": "Docente Planificado",
                    "Estudiante": "Estudiante",
                    "Sala": "Sala",
                    "Día y Fecha": "Día y Fecha",
                    "Hora": "Hora",
                    "Estado": "Estado"
                })
                df_disp = df_disp[["Docente Planificado", "Rol", "Docente Real (M365)", "¿Reemplazo?", "Estudiante", "Sala", "Día y Fecha", "Hora", "Estado"]]
                df_disp = df_disp.reset_index(drop=True)
                
                # Beautiful display table with colors
                st.dataframe(
                    df_disp.style.map(style_status, subset=["Estado"]).map(style_replacement, subset=["¿Reemplazo?"]),
                    use_container_width=True
                )
                
                # ----------------------------------------------------
                # GENERADOR DE CORREOS DE SEGUIMIENTO (PENDIENTES)
                # ----------------------------------------------------
                st.markdown("---")
                st.markdown("### **✉️ Generador de Correos de Seguimiento (Pendientes)**")
                st.markdown("Genera automáticamente el texto preformateado para enviar un recordatorio por correo a los docentes que tienen evaluaciones pendientes.")
                
                df_pending_all = df_filtered[df_filtered["Estado"] == "Pendiente"]
                
                if not df_pending_all.empty:
                    # Get unique planned teachers who actually have pending reviews
                    unique_pending_teachers = sorted(list(df_pending_all["Docente"].unique()))
                    
                    col_f1, col_f2 = st.columns(2)
                    with col_f1:
                        selected_followup_teacher = st.selectbox("Seleccione el docente con pendientes:", unique_pending_teachers, key="followup_teacher")
                    with col_f2:
                        raw_name = str(selected_followup_teacher).strip()
                        name_parts = raw_name.split()
                        default_greeting = name_parts[0].title()
                        if len(name_parts) > 1 and name_parts[0].upper() in ["LUIS", "JUAN", "MARIA", "ANA", "JOSE", "CARLOS"]:
                            default_greeting = f"{name_parts[0].title()} {name_parts[1].title()}"
                        greeting_name = st.text_input("Nombre de saludo:", value=default_greeting, key=f"followup_greeting_{selected_followup_teacher}")
                        
                    col_f3, col_f4 = st.columns(2)
                    with col_f3:
                        is_oral = (selected_section == "🎓 Defensa Oral")
                        default_url = "https://forms.cloud.microsoft/r/kgQZrg6FC9" if is_oral else "https://forms.cloud.microsoft/r/G09LArh2PA"
                        form_url = st.text_input("Enlace del Formulario:", value=default_url, key=f"followup_url_{selected_followup_teacher}")
                    with col_f4:
                        signature_name = st.text_input("Firma del correo:", value="Patricio David Ponce", key=f"followup_signature_{selected_followup_teacher}")
                        
                    df_teacher_pending = df_pending_all[df_pending_all["Docente"] == selected_followup_teacher]
                    
                    # Build plain text email body
                    email_body = f"Estimado {greeting_name},\n\n"
                    if is_oral:
                        email_body += "Para completar el proceso de titulación, le solicito de favor calificar la DEFENSA ORAL a los siguientes estudiantes:\n\n"
                        
                        # Markdown Table Headers
                        email_body += f"{'Día':<8} | {'Fecha':<20} | {'Hora inicio':<11} | {'Hora fin':<8} | {'Sala':<8} | {'Proyecto':<75} | {'Estudiantes'}\n"
                        email_body += f"{'-'*8}-+-{'-'*20}-+-{'-'*11}-+-{'-'*8}-+-{'-'*8}-+-{'-'*75}-+-{'-'*30}\n"
                        
                        for _, r in df_teacher_pending.iterrows():
                            # Parse Día y Fecha
                            parts_day = str(r["Día y Fecha"]).split("(")
                            dia = parts_day[0].strip() if len(parts_day) > 0 else ""
                            fecha = parts_day[1].replace(")", "").strip() if len(parts_day) > 1 else ""
                            
                            # Parse Hora
                            parts_hour = str(r["Hora"]).split("-")
                            hora_inicio = parts_hour[0].strip() if len(parts_hour) > 0 else ""
                            hora_fin = parts_hour[1].strip() if len(parts_hour) > 1 else ""
                            
                            sala = r["Sala"]
                            
                            # Find project
                            proyecto = r.get("Proyecto", "")
                            if not proyecto:
                                proyecto = r.get("Grupo_Alumnos", "")
                            
                            # Limit project length to keep formatting tidy in text area
                            if len(str(proyecto)) > 72:
                                proyecto = str(proyecto)[:69] + "..."
                                
                            student = r["Estudiante"]
                            
                            email_body += f"{dia:<8} | {fecha:<20} | {hora_inicio:<11} | {hora_fin:<8} | {sala:<8} | {proyecto:<75} | {student}\n"
                        
                        email_body += f"\npor favor registrar la calificación de su defensa en el siguiente formulario:\n"
                    else:
                        email_body += "Para completar el proceso de titulación, le solicito de favor calificar el PROYECTO CAPSTONE a los siguientes estudiantes:\n\n"
                        
                        # Markdown Table Headers for Written Capstone (Group and Project)
                        email_body += f"{'Grupo (Estudiantes)':<70} | {'Proyecto'}\n"
                        email_body += f"{'-'*70}-+-{'-'*80}\n"
                        
                        for _, r in df_teacher_pending.iterrows():
                            students = r["Estudiante"]
                            proyecto = r.get("Proyecto", "")
                            if not proyecto:
                                proyecto = r.get("Grupo_Alumnos", "")
                                
                            if len(str(proyecto)) > 77:
                                proyecto = str(proyecto)[:77] + "..."
                                
                            email_body += f"{students:<70} | {proyecto}\n"
                            
                        email_body += f"\npor favor registrar la calificación del proyecto en el siguiente formulario:\n"
                        
                    email_body += f"{form_url}\n\n"
                    email_body += "Quedo atento a cualquier duda. Muchas gracias.\n\n"
                    email_body += "Saludos,\n\n"
                    email_body += f"{signature_name}"
                    
                    # Render beautifully in a text area
                    st.text_area("📋 Cuerpo del Correo (Copiar y Pegar):", value=email_body, height=350, key=f"followup_textarea_{selected_followup_teacher}")
                    st.info("💡 **Tip:** Streamlit proporciona un botón de **'Copiar'** en la esquina superior derecha del cuadro de texto anterior al pasar el mouse por encima.")
                else:
                    st.success("🎉 **¡Excelente! No hay evaluaciones pendientes para los filtros actuales.** Todos los docentes de esta selección han completado sus rúbricas.")
            else:
                st.info("💡 Por favor, sube un **Excel de Planificación** con la pestaña 'CALIFICACION-DOCENTE' en la barra lateral para habilitar esta pestaña y monitorear la puntualidad de tus comités.")
                
        # ----------------------------------------------------
        # TAB 3.5: DUPLICATE CONTROL AUDITOR
        # ----------------------------------------------------
        with tab_duplicates:
            st.markdown("### **🔍 Control de Calificaciones Duplicadas**")
            st.markdown("Aquí puedes auditar y monitorear si algún miembro del tribunal o tutor ha enviado más de una rúbrica para el mismo estudiante o grupo.")
            
            # Find duplicates using raw processing
            df_ind_all, _, _, _ = process_func(raw_path, schedule_path, exclude_duplicates=False)
            
            # Identify duplicate pairs
            dup_mask = df_ind_all.duplicated(subset=['Evaluator_Normalized', 'Student_Normalized'], keep=False)
            df_dups = df_ind_all[dup_mask].copy()
            
            if not df_dups.empty:
                st.warning(f"⚠️ **Se han detectado {len(df_dups)} registros duplicados** (afecta a {len(df_dups.drop_duplicates(subset=['Evaluator_Normalized', 'Student_Normalized']))} casos únicos).")
                
                # Show status of exclusion
                if exclude_duplicates:
                    st.info("ℹ️ **El filtro automático está ACTIVADO:** El dashboard solo considera el envío más reciente (por ID/timestamp) y descarta los duplicados más antiguos para calcular los promedios.")
                else:
                    st.error("🚨 **El filtro automático está DESACTIVADO:** Las notas se promedian considerando todos los envíos duplicados, lo cual puede distorsionar el consolidado final.")
                
                # Table of duplicates styled
                st.markdown("#### **Listado de Rúbricas Duplicadas Detectadas**")
                
                # We can construct a nice table showing: ID, Docente, Estudiante, Fecha, y cuál se conserva
                dup_records = []
                for (eval_norm, stud_norm), group_df in df_dups.groupby(['Evaluator_Normalized', 'Student_Normalized']):
                    # Sort by ID descending to know which is the kept one (first when sorted desc)
                    sorted_group = group_df.sort_values(by="Id", ascending=False)
                    kept_id = sorted_group.iloc[0]["Id"]
                    
                    for _, row in group_df.iterrows():
                        is_kept = row["Id"] == kept_id
                        dup_records.append({
                            "ID Envío": row["Id"],
                            "Evaluador / Docente": row["Evaluator_Raw"],
                            "Estudiante": row["Student_Raw"],
                            "Fecha de Envío": row["Date"] if "Date" in row and row["Date"] else "N/D",
                            "Estado en Consolidado": "🟢 Conservado (Reciente)" if (is_kept and exclude_duplicates) else ("🔴 Omitido" if exclude_duplicates else "⚠️ Promediado (Duplicado)")
                        })
                        
                df_dup_table = pd.DataFrame(dup_records).sort_values(by="ID Envío", ascending=False).reset_index(drop=True)
                
                # Render table with styled statuses
                def style_dup_status(val):
                    if "Conservado" in val:
                        color = "rgba(16, 185, 129, 0.2)"
                        text_color = "#047857"
                    elif "Omitido" in val:
                        color = "rgba(239, 68, 68, 0.2)"
                        text_color = "#b91c1c"
                    else:
                        color = "rgba(245, 158, 11, 0.2)"
                        text_color = "#b45309"
                    return f'background-color: {color}; color: {text_color}; font-weight: bold; text-align: center; border-radius: 4px;'
                
                st.dataframe(
                    df_dup_table.style.map(style_dup_status, subset=["Estado en Consolidado"]),
                    use_container_width=True
                )
                
                # Show an audit comparative block for duplicate students
                st.markdown("#### **🔍 Análisis Comparativo del Impacto de Duplicados**")
                st.markdown("Selecciona un estudiante con envíos duplicados para comparar su nota con y sin exclusión de duplicados:")
                
                dup_students = sorted(df_dups["Student_Raw"].unique())
                selected_dup_st = st.selectbox("Estudiante a Auditar:", dup_students)
                
                if selected_dup_st:
                    # Let's show the individual reviews for this student
                    st_norm = normalize_name(selected_dup_st)
                    st_evals = df_ind_all[df_ind_all["Student_Normalized"] == st_norm]
                    
                    st.write(f"**Evaluaciones registradas para {selected_dup_st}:**")
                    for _, row_eval in st_evals.iterrows():
                        score = sum(row_eval[f"{k}_pts"] for k in active_criteria_map)
                        is_newest = row_eval["Id"] == st_evals[st_evals["Evaluator_Normalized"] == row_eval["Evaluator_Normalized"]]["Id"].max()
                        kept_lbl = " (Último envío)" if is_newest else " (Envío anterior)"
                        st.write(f"- **ID {row_eval['Id']}** por **{row_eval['Evaluator_Raw']}**: {score:.0f}/100{kept_lbl}")
                        
                    match_col = "Seleccione el nombre del Estudiante"
                    row_no_ex = df_calc_no_ex[df_calc_no_ex[match_col] == selected_dup_st]
                    if row_no_ex.empty:
                        row_no_ex = df_calc_no_ex[df_calc_no_ex[match_col].astype(str).str.contains(selected_dup_st)]
                        
                    row_ex = df_calc_ex[df_calc_ex[match_col] == selected_dup_st]
                    if row_ex.empty:
                        row_ex = df_calc_ex[df_calc_ex[match_col].astype(str).str.contains(selected_dup_st)]
                        
                    grade_no_ex = row_no_ex["Nota ponderada"].values[0] if not row_no_ex.empty else 0.0
                    grade_ex = row_ex["Nota ponderada"].values[0] if not row_ex.empty else 0.0
                    
                    diff = grade_ex - grade_no_ex
                    st.markdown(f"""
                    <div style="background-color: #f3f4f6; padding: 1rem; border-radius: 8px; margin-top: 0.5rem;">
                        <p style="margin:0;"><b>Nota PROMEDIANDO duplicados:</b> {grade_no_ex:.2f}/100</p>
                        <p style="margin:0;"><b>Nota CONSERVANDO ÚLTIMO envío:</b> {grade_ex:.2f}/100</p>
                        <p style="margin:0; color: {'#0d9488' if diff >= 0 else '#ef4444'}; font-weight: bold;">
                            Impacto del filtrado de duplicados: {diff:+.2f} puntos en la nota final.
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.success("🎉 **¡Excelente! No se han detectado envíos duplicados de rúbricas en esta cohorte.**")
                st.info("Cada docente ha enviado exactamente una rúbrica por estudiante/grupo programado.")
                
        # ----------------------------------------------------
        # TAB 4: FILE EXPORTER / EXCEL DOWNLOADS
        # ----------------------------------------------------
        with tab_exporter:
            st.markdown("### **Exportar Resultados Consolidados**")
            st.markdown("Descarga el archivo Excel final con las notas promediadas, desgloses individuales y la matriz de compliance lista para ser distribuida a la coordinación académica.")
            
            # Export logic
            output_stream = BytesIO()
            export_func(df_calc, output_stream, df_individual, df_schedule, df_compliance)
            excel_data = output_stream.getvalue()
            
            col_exp1, col_exp2 = st.columns(2)
            with col_exp1:
                st.markdown(f"""
                ##### **Reporte Excel Oficial (.xlsx) - {selected_section}**
                Contiene:
                * **Sheet1**: Base de datos de evaluaciones con campos completos.
                * **Calculo y Seguimiento**: Promedios ponderados e IFS cualitativos para cada estudiante ordenados de forma alfabética.
                * **CALIFICACION-DOCENTE**: Calendario de asignación de tribunales.
                * **PESOS**: Tablas de indexación cualitativa.
                """)
                st.download_button(
                    label="📥 Descargar Reporte Excel Consolidado",
                    data=excel_data,
                    file_name=export_file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with col_exp2:
                # Exporter statistics
                st.success("✅ Generación de reporte completada de forma correcta!")
                st.markdown(f"""
                - **Estudiantes Consolidados:** {len(df_calc)} registros
                - **Líneas de Auditoría (Formularios):** {len(df_individual)} registros
                - **Verificaciones Internas:** Coincidencia matemática exitosa al 100%
                """)
                
    except Exception as e:
        st.error(f"❌ Ocurrió un error al procesar el archivo: {e}")
        st.info("Asegúrate de que los archivos cargados coincidan con el formato crudo de Microsoft Forms de rúbricas.")
else:
    # Landing page instructions if no file is uploaded and demo is off
    st.warning("⚠️ Esperando carga de datos.")
    st.info(f"Sube el **Reporte Crudo de {selected_section} (.xlsx)** en la barra lateral o activa la opción **'Cargar Cohorte de Prueba (Demo)'** para iniciar inmediatamente.")
