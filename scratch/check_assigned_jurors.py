import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import openpyxl
from collections import Counter
from src.cleaner import normalize_name

wb = openpyxl.load_workbook("data/presentaciones_crcronograma.xlsx", data_only=True)
sheet = wb["Hoja1"]
headers = [cell.value for cell in sheet[1]]

idx_name = headers.index("NOMBRE")
idx_tit = headers.index("DOCENTE TITULACIÓN")
idx_tutor = headers.index("TUTOR")
idx_tercer = headers.index("TERCER DOCENTE")
idx_adic = headers.index("DOCENTE ADICIONAL")

evals_assigned_counts = Counter()
for r in range(2, sheet.max_row+1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, len(headers)+1)]
    if any(vals):
        name = vals[idx_name]
        if name:
            assigned = []
            for col_idx in [idx_tit, idx_tutor, idx_tercer, idx_adic]:
                val = vals[col_idx]
                if val and str(val).strip() != "" and str(val).strip().lower() not in ["none", "nan", "null"]:
                    assigned.append(val)
            evals_assigned_counts[len(assigned)] += 1
            if len(assigned) != 3:
                print(f"Student: {name} has {len(assigned)} assigned jurors: {assigned}")

print("\nDistribution of assigned jurors count per student:")
print(dict(evals_assigned_counts))
