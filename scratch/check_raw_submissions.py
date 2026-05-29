import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import openpyxl
from collections import Counter
from src.cleaner import normalize_name

wb = openpyxl.load_workbook("data/DEFENSA ORAL DE PROYECTO CAPSTONE - COHORTE 2(1-360).xlsx", data_only=True)
sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
headers = [c.value for c in sheet[1]]

student_col_idx = None
for idx, h in enumerate(headers):
    if h and "Estudiante" in str(h):
        student_col_idx = idx
        break

counts = Counter()
for r in range(2, sheet.max_row+1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, len(headers)+1)]
    if any(vals):
        s_name = vals[student_col_idx]
        if s_name and str(s_name).strip() != "" and str(s_name).strip().lower() not in ["none", "nan"]:
            counts[normalize_name(s_name)] += 1

print("Distribution of submissions per student in Cohorte 2 raw file:")
print(dict(Counter(counts.values())))

for s, c in counts.items():
    if c != 3:
        print(f"Student {s} has {c} submissions")
