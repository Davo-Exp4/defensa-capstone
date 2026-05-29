import openpyxl
import pandas as pd

wb = openpyxl.load_workbook("data/presentaciones_crcronograma.xlsx", data_only=True)
print("Sheetnames:", wb.sheetnames)
if "Hoja1" in wb.sheetnames:
    sheet = wb["Hoja1"]
    headers = [cell.value for cell in sheet[1]]
    print("Hoja1 Headers:", headers)
    
    rows = []
    for r in range(2, sheet.max_row+1):
        row_vals = [cell.value for cell in sheet[r]]
        if any(row_vals):
            rows.append(row_vals)
    df = pd.DataFrame(rows, columns=headers)
    print("Shape:", df.shape)
    print("Unique student count in Hoja1 NOMBRE column:", df[df.columns[1]].nunique())
    print("Unique student count normalized NOMBRE column:", df[df.columns[1]].apply(lambda x: str(x).strip().upper()).nunique())
    
    print("\nColumns check:")
    for col in ["DOCENTE TITULACIÓN", "TUTOR", "TERCER DOCENTE", "DOCENTE ADICIONAL"]:
        if col in df.columns:
            print(f"Unique {col}: {df[col].dropna().unique()}")
            
if "Proyectos agrupados" in wb.sheetnames:
    sheet_pa = wb["Proyectos agrupados"]
    headers_pa = [cell.value for cell in sheet_pa[1]]
    print("\nProyectos agrupados Headers:", headers_pa)
