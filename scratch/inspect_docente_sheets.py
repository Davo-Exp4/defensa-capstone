import openpyxl
import pandas as pd

wb = openpyxl.load_workbook("data/presentaciones_crcronograma.xlsx", data_only=True)
for sheetname in ["Resumen docentes", "Conexión docentes"]:
    if sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        print(f"\n=================== Sheet: {sheetname} ===================")
        headers = [cell.value for cell in sheet[1]]
        print("Headers:", headers)
        
        rows = []
        for r in range(2, min(sheet.max_row+1, 10)):
            row_vals = [cell.value for cell in sheet[r]]
            if any(row_vals):
                rows.append(row_vals)
        df = pd.DataFrame(rows, columns=headers[:len(row_vals)])
        print("First few rows:")
        print(df.to_string())
