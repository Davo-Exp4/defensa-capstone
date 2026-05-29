import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import openpyxl
from src.cleaner import normalize_name
from src.engine import process_oral_defense

df_ind, df_calc, df_comp, _ = process_oral_defense(
    "data/DEFENSA ORAL DE PROYECTO CAPSTONE - COHORTE 2(1-360).xlsx",
    "data/presentaciones_crcronograma.xlsx",
    exclude_duplicates=True
)

# Let's count how many scheduled rows we have
print("Current compliance records total:", len(df_comp))
print("Current completed records total:", len(df_comp[df_comp["Estado"] == "Completado"]))

# Let's see what happens if we deduplicate the assigned jurors for each student in Hoja1
wb_sched = openpyxl.load_workbook("data/presentaciones_crcronograma.xlsx", data_only=True)
sheet_sched = wb_sched["Hoja1"]
headers_sched = [cell.value for cell in sheet_sched[1]]

idx_name = headers_sched.index("NOMBRE")
idx_tit = headers_sched.index("DOCENTE TITULACIÓN")
idx_tutor = headers_sched.index("TUTOR")
idx_tercer = headers_sched.index("TERCER DOCENTE")
idx_adic = headers_sched.index("DOCENTE ADICIONAL")

total_planned_dedup = 0
completed_dedup = 0

for r in range(2, sheet_sched.max_row+1):
    vals = [sheet_sched.cell(row=r, column=c).value for c in range(1, len(headers_sched)+1)]
    if any(vals):
        student_raw = vals[idx_name]
        if student_raw and str(student_raw).strip() != "" and str(student_raw).strip().lower() not in ["none", "nan"]:
            student_norm = normalize_name(student_raw)
            
            # Extract assigned jurors and preserve their roles while deduplicating by normalized name
            seen_jurors = {}
            for col_idx, role in [(idx_tit, "Docente titulación"), (idx_tutor, "Tutor"), (idx_tercer, "Tercer docente"), (idx_adic, "Docente adicional")]:
                juror_raw = vals[col_idx]
                if juror_raw and str(juror_raw).strip() != "" and str(juror_raw).strip().lower() not in ["none", "nan"]:
                    juror_norm = normalize_name(juror_raw)
                    if juror_norm not in seen_jurors:
                        seen_jurors[juror_norm] = (juror_raw, role)
                    else:
                        # If already seen, we can combine roles or keep the first one
                        pass
            
            total_planned_dedup += len(seen_jurors)
            
            # Check submission for each unique juror
            for juror_norm, (juror_raw, role) in seen_jurors.items():
                submitted = df_ind[
                    (df_ind["Student_Normalized"] == student_norm) &
                    (df_ind["Evaluator_Normalized"] == juror_norm)
                ]
                if not submitted.empty:
                    completed_dedup += 1

print("\n--- Deduplicated Jurors Math ---")
print("Total Unique Planned Evaluations (132 * 3):", total_planned_dedup)
print("Completed Unique Evaluations:", completed_dedup)
print("Remaining Unique Evaluations:", total_planned_dedup - completed_dedup)
print("Percentage Done:", (completed_dedup / total_planned_dedup) * 100)
