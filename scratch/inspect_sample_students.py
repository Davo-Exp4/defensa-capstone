import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import openpyxl
from src.cleaner import normalize_name, split_group_names

wb_raw = openpyxl.load_workbook("data/DEFENSA ORAL DE PROYECTO CAPSTONE - COHORTE 2(1-360).xlsx", data_only=True)
sheet_raw = wb_raw.active
headers_raw = [cell.value for cell in sheet_raw[1]]

student_col = None
evaluator_col = None
submitter_col = 4 # Email is col 3, Name is col 4 (0-indexed)
id_col = 0

for idx, h in enumerate(headers_raw):
    if h and "Estudiante" in str(h):
        student_col = idx
    if h and "Evaluador" in str(h):
        evaluator_col = idx

print("Student col:", student_col, headers_raw[student_col])
print("Evaluator col:", evaluator_col, headers_raw[evaluator_col])

wb_sched = openpyxl.load_workbook("data/presentaciones_crcronograma.xlsx", data_only=True)
sheet_sched = wb_sched["Hoja1"]
headers_sched = [cell.value for cell in sheet_sched[1]]

def get_scheduled_jurors(student_name):
    norm_name = normalize_name(student_name)
    idx_name = headers_sched.index("NOMBRE")
    idx_tit = headers_sched.index("DOCENTE TITULACIÓN")
    idx_tutor = headers_sched.index("TUTOR")
    idx_tercer = headers_sched.index("TERCER DOCENTE")
    idx_adic = headers_sched.index("DOCENTE ADICIONAL")
    
    for r in range(2, sheet_sched.max_row+1):
        vals = [sheet_sched.cell(row=r, column=c).value for c in range(1, len(headers_sched)+1)]
        if any(vals):
            if normalize_name(vals[idx_name]) == norm_name:
                jurors = []
                for idx in [idx_tit, idx_tutor, idx_tercer, idx_adic]:
                    v = vals[idx]
                    if v and str(v).strip() != "" and str(v).strip().lower() not in ["none", "nan"]:
                        jurors.append(v)
                return jurors
    return []

# Inspect a few 2-eval and 4-eval students
sample_students = [
    "RAMOS VASCONEZ XAVIER ALEJANDRO", # 2 evals
    "AGUILAR RODRIGUEZ CAMILA ESTEFANIA", # 4 evals
    "CHIRIBOGA TERAN JUAN MARTIN" # 4 evals
]

for s in sample_students:
    print(f"\n==================== {s} ====================")
    sched = get_scheduled_jurors(s)
    print("Scheduled Jurors:", sched)
    
    # Submissions
    norm_s = normalize_name(s)
    print("Submissions in raw:")
    for r in range(2, sheet_raw.max_row+1):
        vals = [sheet_raw.cell(row=r, column=c).value for c in range(1, len(headers_raw)+1)]
        if any(vals):
            student_raw = vals[student_col]
            if student_raw and normalize_name(student_raw) == norm_s:
                id_val = vals[id_col]
                evaluator_raw = vals[evaluator_col]
                email_val = vals[3]
                submitter_val = vals[4]
                print(f"  ID: {id_val} | Selected Evaluator: {evaluator_raw} | Submitter M365 Name: {submitter_val} | Submitter Email: {email_val}")
