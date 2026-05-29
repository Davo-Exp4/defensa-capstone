import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import pandas as pd
import openpyxl
from src.engine import process_oral_defense, CRITERIA_MAP

def run_partial_test():
    raw_full_path = "data/DEFENSA ORAL DE PROYECTO CAPSTONE - COHORTE 2(1-360).xlsx"
    processed_path = "data/DEFENSA ORAL DE PROYECTO CAPSTONE - COHORTE 2_PROCESADO.xlsx"
    schedule_path = "data/presentaciones_crcronograma.xlsx"
    
    # Load 360 raw workbook, and save only first 95 rows (including header) to a temp file
    wb = openpyxl.load_workbook(raw_full_path)
    sheet = wb.active
    
    # Delete rows after 95
    if sheet.max_row > 95:
        sheet.delete_rows(96, sheet.max_row - 95)
        
    temp_raw_path = "scratch/temp_raw_94.xlsx"
    wb.save(temp_raw_path)
    print(f"Saved partial raw workbook to {temp_raw_path} with {sheet.max_row} rows.")
    
    # Process using our engine
    df_ind, df_calc, df_comp, _ = process_oral_defense(temp_raw_path, schedule_path, exclude_duplicates=False)
    df_calc.columns = [str(c).strip() for c in df_calc.columns]
    
    # Load processed Excel Calculo sheet
    df_calc_hist = pd.read_excel(processed_path, sheet_name="Calculo")
    df_calc_hist.columns = [str(c).strip() for c in df_calc_hist.columns]
    df_calc_hist = df_calc_hist[df_calc_hist["Seleccione el nombre del Estudiante"].notna()]
    df_calc_hist = df_calc_hist[df_calc_hist["Seleccione el nombre del Estudiante"].str.upper() != "TOTAL GENERAL"]
    
    df_calc["student_key"] = df_calc["Seleccione el nombre del Estudiante"].str.strip().str.upper()
    df_calc_hist["student_key"] = df_calc_hist["Seleccione el nombre del Estudiante"].str.strip().str.upper()
    
    ours_indexed = df_calc.set_index("student_key")
    hist_indexed = df_calc_hist.set_index("student_key")
    
    print("\n--- Verifying first 94 evaluations ---")
    mismatches = 0
    checked_students = 0
    
    columns_to_compare = [
        "Cuenta de Seleccione su nombre (Evaluador)",
        "Nota ponderada",
    ]
    for key, c_info in CRITERIA_MAP.items():
        columns_to_compare.append(c_info["clean_name"].strip())
        columns_to_compare.append(c_info["qual_name"].strip())
        
    for s_key in ours_indexed.index:
        if s_key in hist_indexed.index:
            checked_students += 1
            row_ours = ours_indexed.loc[s_key]
            row_hist = hist_indexed.loc[s_key]
            student_name = row_ours["Seleccione el nombre del Estudiante"]
            
            for col in columns_to_compare:
                if col not in row_ours or col not in row_hist:
                    continue
                val_ours = row_ours[col]
                val_hist = row_hist[col]
                
                if isinstance(val_ours, (int, float)) and isinstance(val_hist, (int, float)):
                    if abs(val_ours - val_hist) > 1e-5:
                        print(f"   [ERROR] Mismatch para {student_name} en '{col}': Nuestro={val_ours:.4f}, Histórico={val_hist:.4f}")
                        mismatches += 1
                else:
                    if str(val_ours).strip().lower() != str(val_hist).strip().lower():
                        print(f"   [ERROR] Mismatch para {student_name} en '{col}': Nuestro='{val_ours}', Histórico='{val_hist}'")
                        mismatches += 1
                        
    print(f"Partial test completed. Students checked: {checked_students}, Mismatches: {mismatches}")
    if mismatches == 0:
        print("🎉 ¡EXITO ROTUNDO! Con los mismos datos de entrada, la lógica de cálculo del backend es 100% IDÉNTICA al histórico.")

if __name__ == "__main__":
    run_partial_test()
